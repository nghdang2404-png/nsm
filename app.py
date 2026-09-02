import os
import re
import io
import json
import unicodedata
import ssl
import threading
import time
import traceback
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, has_request_context, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import case, text
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from PIL import Image, ImageOps

ssl._create_default_https_context = ssl._create_unverified_context

app = Flask(__name__)
# Ưu tiên đọc SECRET_KEY từ biến môi trường (set trong Render > Environment).
# Nếu chưa set thì fallback về giá trị cũ để chạy local không bị lỗi.
app.secret_key = os.environ.get('SECRET_KEY', 'super_secret_key')

# Render tự động set biến môi trường RENDER=true trên server production.
# Dùng để bật các cấu hình chỉ nên áp dụng khi chạy thật (HTTPS), không áp dụng khi
# chạy local qua http://127.0.0.1 (nếu bật Secure=True lúc chạy local, cookie sẽ
# không được set vì không có HTTPS -> không đăng nhập được).
IS_PRODUCTION = os.environ.get('RENDER') == 'true'

# --- CẤU HÌNH COOKIE ĐỂ GIỮ PHIÊN TRÊN ĐIỆN THOẠI ---
app.permanent_session_lifetime = timedelta(days=30)
app.config['SESSION_COOKIE_NAME'] = 'namsuong_session'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = IS_PRODUCTION

UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# --- CACHE ẢNH TĨNH (logo, ảnh xe theo màu...) TRÊN TRÌNH DUYỆT ---
# Mặc định Flask không set Cache-Control cho /static, nên mỗi lần đổi màu xe (kể cả
# đã xem màu đó trước đó) trình duyệt vẫn phải hỏi lại server -> chậm.
# Set 30 ngày để trình duyệt tự phục vụ từ cache cục bộ cho các lần sau, ảnh chỉ tải
# lại khi tên file thay đổi (khi admin thay ảnh mới, tên file thường đổi hoặc cần thêm
# ?v=... nếu ghi đè cùng tên file).
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 60 * 60 * 24 * 30

# Ưu tiên đọc chuỗi kết nối từ biến môi trường DATABASE_URL (set trong Render >
# Environment). Fallback về chuỗi cũ để vẫn chạy được khi test local mà chưa set env.
# LƯU Ý: nên đổi mật khẩu DB này vì đã từng bị lộ thẳng trong code nguồn.
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL',
    'postgresql://neondb_owner:npg_knMXRhS06HbT@ep-fancy-block-az7pz4uf.c-3.ap-southeast-1.aws.neon.tech/neondb?sslmode=require&connect_timeout=30'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True, 'pool_recycle': 300}

db = SQLAlchemy(app)

# --- HÀM TỰ ĐỘNG PHÂN LOẠI DỰA TRÊN TỪ KHÓA TÊN XE ---
def tu_dong_phan_loai(ten_xe):
    if not ten_xe:
        return "Chưa phân loại"
    
    t = unicodedata.normalize('NFC', str(ten_xe)).lower()
    
    if any(kw in t for kw in ['airblade', 'air blade', 'vision', 'lead', 'sh mode', 'sh', 'vario', 'scoopy', 'pcx']):
        return "Xe ga"
    elif any(kw in t for kw in ['wave', 'blade', 'future', 'super cub', 'dream']):
        return "Xe số"
    elif any(kw in t for kw in ['winner', 'cb', 'cbr', 'rebel', 'sonic', 'côn tay']):
        return "Xe côn tay & Thể thao"
        
    return "Chưa phân loại"

# --- MODELS ---
class Setting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.Text, nullable=True)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    ho_ten = db.Column(db.String(150))
    role = db.Column(db.String(50), default='user')
    bo_phan = db.Column(db.String(100))
    khu_vuc = db.Column(db.String(100))
    trang_thai = db.Column(db.String(20), default='approved') # 'pending' (chờ duyệt), 'approved' (đã duyệt), 'rejected' (từ chối)

class Xe(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    loai_xe = db.Column(db.String(50))
    ten_xe = db.Column(db.String(100), unique=True, nullable=False)
    phien_ban = db.Column(db.String(100))
    
    gia_cm_thap = db.Column(db.Float, default=0)
    gia_cm_trung = db.Column(db.Float, default=0)
    gia_cm_cao = db.Column(db.Float, default=0)
    
    gia_bl_thap = db.Column(db.Float, default=0)
    gia_bl_trung = db.Column(db.Float, default=0)
    gia_bl_cao = db.Column(db.Float, default=0)

    gia_gt_phuong_cm = db.Column(db.Float, default=0)
    gia_gt_xa_cm = db.Column(db.Float, default=0)
    gia_gt_phuong_bl = db.Column(db.Float, default=0)
    gia_gt_xa_bl = db.Column(db.Float, default=0)
    
    ns1 = db.Column(db.Integer, default=0); ns2 = db.Column(db.Integer, default=0)
    ns3 = db.Column(db.Integer, default=0); ns4 = db.Column(db.Integer, default=0)
    ns5 = db.Column(db.Integer, default=0); nsm1 = db.Column(db.Integer, default=0)
    hinh_anh = db.Column(db.String(200), default='')
    mau_xe = db.relationship('XeMau', backref='xe', cascade='all, delete-orphan')

class XeMau(db.Model):
    __tablename__ = 'xe_mau'
    id = db.Column(db.Integer, primary_key=True)
    xe_id = db.Column(db.Integer, db.ForeignKey('xe.id'), nullable=False)
    ten_mau = db.Column(db.String(50), nullable=False)
    chenh_lech_cm = db.Column(db.Float, default=0)
    chenh_lech_bl = db.Column(db.Float, default=0)
    hinh_anh_mau = db.Column(db.String(200), default='')

    ns1 = db.Column(db.Integer, default=0)
    ns2 = db.Column(db.Integer, default=0)
    ns3 = db.Column(db.Integer, default=0)
    ns4 = db.Column(db.Integer, default=0)
    ns5 = db.Column(db.Integer, default=0)
    nsm1 = db.Column(db.Integer, default=0)

    def to_dict(self, khu_vuc_user=None):
        is_bl = 'bạc liêu' in (khu_vuc_user or '').lower()
        chenh_lech_vung = self.chenh_lech_bl if is_bl else self.chenh_lech_cm
        return {
            'ten_mau': self.ten_mau, 
            'chenh_lech_gia': chenh_lech_vung,
            'chenh_lech_cm': self.chenh_lech_cm or 0,
            'chenh_lech_bl': self.chenh_lech_bl or 0,
            'ds_ma_mau': lay_danh_sach_ma_mau(self.ten_mau),
            'hinh_anh_mau': self.hinh_anh_mau,
            'ns1': self.ns1 or 0,
            'ns2': self.ns2 or 0,
            'ns3': self.ns3 or 0,
            'ns4': self.ns4 or 0,
            'ns5': self.ns5 or 0,
            'nsm1': self.nsm1 or 0
        }

class KhuVucLonBL(db.Model):
    """Khu vực lớn về giấy tờ ở Bạc Liêu: Nam Sương 4, Nam Sương 2, Nam Sương 5 - NSM1."""
    __tablename__ = 'khu_vuc_lon_bl'
    id = db.Column(db.Integer, primary_key=True)
    ma_khu_vuc = db.Column(db.String(20), unique=True, nullable=False)   # 'NS4', 'NS2', 'NSM1'
    ten_khu_vuc = db.Column(db.String(100), nullable=False)              # 'Nam Sương 4'
    thu_tu = db.Column(db.Integer, default=0)
    khu_vuc_nho = db.relationship(
        'KhuVucNhoBL', backref='khu_vuc_lon',
        cascade='all, delete-orphan', order_by='KhuVucNhoBL.thu_tu'
    )

class KhuVucNhoBL(db.Model):
    """Khu vực nhỏ (nhóm phường/xã đồng mức giá) bên trong 1 khu vực lớn Bạc Liêu."""
    __tablename__ = 'khu_vuc_nho_bl'
    id = db.Column(db.Integer, primary_key=True)
    khu_vuc_lon_id = db.Column(db.Integer, db.ForeignKey('khu_vuc_lon_bl.id'), nullable=False)
    ten_khu_vuc_nho = db.Column(db.String(255), nullable=False)          # "Phường Giá Rai, Phường Láng Tròn, Xã Phong Thạnh"
    thu_tu = db.Column(db.Integer, default=0)

class GiaGiayToXeBL(db.Model):
    """Giá giấy tờ của 1 xe tại 1 khu vực nhỏ thuộc Bạc Liêu (mỗi xe có thể có giá giấy tờ khác nhau theo từng khu vực nhỏ)."""
    __tablename__ = 'gia_giay_to_xe_bl'
    id = db.Column(db.Integer, primary_key=True)
    xe_id = db.Column(db.Integer, db.ForeignKey('xe.id'), nullable=False)
    khu_vuc_nho_id = db.Column(db.Integer, db.ForeignKey('khu_vuc_nho_bl.id'), nullable=False)
    gia = db.Column(db.Float, default=0)

    khu_vuc_nho = db.relationship('KhuVucNhoBL')

    __table_args__ = (db.UniqueConstraint('xe_id', 'khu_vuc_nho_id', name='uq_xe_khuvucnho_bl'),)

# --- DỮ LIỆU MẶC ĐỊNH CÁC KHU VỰC GIẤY TỜ BẠC LIÊU ---
DU_LIEU_KHU_VUC_GIAY_TO_BL = [
    {
        'ma_khu_vuc': 'NS4',
        'ten_khu_vuc': 'Nam Sương 4',
        'khu_vuc_nho': [
            'Phường Giá Rai, Phường Láng Tròn, Xã Phong Thạnh',
            'Xã Đông Hải, Xã Gành Hào, Xã Định Thành, Xã Long Điền',
            'Xã Phước Long, Xã Vĩnh Phước, Xã Vĩnh Thanh',
            'Xã An Trạch',
            'Xã Phong Hiệp',
        ]
    },
    {
        'ma_khu_vuc': 'NS2',
        'ten_khu_vuc': 'Nam Sương 2',
        'khu_vuc_nho': [
            'Phường Bạc Liêu, Phường Hiệp Thành, Phường Vĩnh Trạch',
            'Vĩnh Lợi, Hoà Bình',
            'Phước Long, Hồng Dân',
            'Đông Hải',
            'Sóc Trăng, Cần Thơ',
        ]
    },
    {
        'ma_khu_vuc': 'NSM1',
        'ten_khu_vuc': 'Nam Sương 5 - NSM1',
        'khu_vuc_nho': [
            'Phường Bạc Liêu, Vĩnh Trạch, Hiệp Thành',
            'Hoà Bình, Vĩnh Lợi',
            'Phước Long',
            'Đông Hải',
            'Giá Rai',
            'Hồng Dân',
            'Sóc Trăng, Cần Thơ',
        ]
    },
]

class KhuyenMai(db.Model):
    """Chương trình khuyến mãi áp dụng cho các dòng xe (khuyến mãi của công ty hoặc của Honda)."""
    __tablename__ = 'khuyen_mai'
    id = db.Column(db.Integer, primary_key=True)
    loai = db.Column(db.String(20), nullable=False)  # 'cty' hoặc 'honda'
    tieu_de = db.Column(db.String(200), nullable=False)
    noi_dung = db.Column(db.Text, default='')
    ngay_bat_dau = db.Column(db.Date, nullable=False)
    ngay_ket_thuc = db.Column(db.Date, nullable=False)
    dong_xe_json = db.Column(db.Text, default='[]')  # danh sách dòng xe ngắn gọn, VD: ["Vision","Wave"]
    dang_bat = db.Column(db.Boolean, default=True)  # cho phép admin tắt thủ công dù còn hạn
    created_at = db.Column(db.DateTime, default=lambda: datetime.now())
    created_by = db.Column(db.String(80))

    @property
    def danh_sach_dong_xe(self):
        try:
            return json.loads(self.dong_xe_json or '[]')
        except Exception:
            return []

    @danh_sach_dong_xe.setter
    def danh_sach_dong_xe(self, ds):
        self.dong_xe_json = json.dumps(ds or [], ensure_ascii=False)

    def dang_hoat_dong(self):
        """CTKM được coi là 'đang hoạt động' khi hôm nay nằm trong khoảng ngày bắt đầu - kết thúc
        và admin chưa tắt thủ công. Tự động hiển thị/ẩn hoàn toàn dựa vào ngày, không cần thao tác thủ công."""
        if not self.dang_bat:
            return False
        homnay = datetime.now().date()
        return self.ngay_bat_dau <= homnay <= self.ngay_ket_thuc

    def trang_thai_text(self):
        homnay = datetime.now().date()
        if not self.dang_bat:
            return 'Đã tắt'
        if homnay < self.ngay_bat_dau:
            return 'Sắp diễn ra'
        if homnay > self.ngay_ket_thuc:
            return 'Đã kết thúc'
        return 'Đang diễn ra'

    def to_dict(self):
        return {
            'id': self.id,
            'loai': self.loai,
            'tieu_de': self.tieu_de,
            'noi_dung': self.noi_dung or '',
            'ngay_bat_dau': self.ngay_bat_dau.strftime('%d/%m/%Y') if self.ngay_bat_dau else '',
            'ngay_ket_thuc': self.ngay_ket_thuc.strftime('%d/%m/%Y') if self.ngay_ket_thuc else '',
        }

# --- DANH SÁCH TỪ KHÓA DÒNG XE (dùng để chọn nhanh dòng xe áp dụng khuyến mãi) ---
# Thứ tự quan trọng: từ khóa dài/cụ thể hơn phải đứng trước để so khớp đúng
# (VD: "Air Blade" phải đứng trước "Blade", "SH Mode" phải đứng trước "SH").
DANH_SACH_TU_KHOA_DONG_XE = [
    'Air Blade 125', 'Air Blade 160', 'Air Blade',
    'SH Mode', 'SH 125', 'SH 160', 'SH',
    'Vision', 'Lead', 'Vario', 'Scoopy', 'PCX',
    'Wave', 'Blade', 'Future', 'Super Cub', 'Dream',
    'Winner', 'CBR', 'CB', 'Rebel', 'Sonic', 'Monkey', 'MSX'
]

def _chuan_hoa_de_so_khop(s):
    """Chuẩn hoá chuỗi để so khớp dòng xe: bỏ dấu cách/gạch ngang/gạch dưới/chấm, viết thường.
    Nhờ vậy 'Air Blade', 'AirBlade', 'Air-Blade' đều được coi là khớp nhau, tránh bị sót
    dòng xe chỉ vì cách viết tên xe trong CSDL không có khoảng trắng giống hệt từ khóa."""
    s = unicodedata.normalize('NFC', str(s or '')).lower()
    return re.sub(r'[\s\-_.]+', '', s)

def lay_danh_sach_dong_xe_hien_co():
    """Quét toàn bộ tên xe hiện có trong CSDL, chỉ trả về những dòng xe THỰC SỰ đang tồn tại
    để hiển thị lên danh sách tick chọn cho admin. Mỗi xe chỉ được xếp vào ĐÚNG 1 dòng xe
    cụ thể nhất (ưu tiên từ khóa dài/cụ thể hơn), tránh liệt kê dư thừa những dòng xe
    'cha' không thực sự có mặt riêng biệt (VD: CSDL chỉ có Air Blade thì KHÔNG hiện
    thêm dòng 'Blade', dù chữ 'Blade' có nằm trong tên 'Air Blade')."""
    danh_sach_ten = [t[0] for t in db.session.query(Xe.ten_xe).all()]
    dong_xe_thuc_te = set()
    for ten in danh_sach_ten:
        dx = _xac_dinh_dong_xe_cua_ten(ten)
        if dx:
            dong_xe_thuc_te.add(dx)
    # Giữ đúng thứ tự ưu tiên đã khai báo trong DANH_SACH_TU_KHOA_DONG_XE khi hiển thị
    return [tu_khoa for tu_khoa in DANH_SACH_TU_KHOA_DONG_XE if tu_khoa in dong_xe_thuc_te]

def _xac_dinh_dong_xe_cua_ten(ten_xe):
    """Xác định DUY NHẤT 1 dòng xe (ngắn gọn) mà 1 chiếc xe thuộc về, dựa vào từ khóa khớp
    ĐẦU TIÊN trong DANH_SACH_TU_KHOA_DONG_XE (từ khóa dài/cụ thể hơn được ưu tiên trước,
    VD: 'Air Blade' được xét trước 'Blade', 'SH Mode' được xét trước 'SH').
    Trả về None nếu tên xe không khớp dòng xe nào."""
    t = _chuan_hoa_de_so_khop(ten_xe)
    for tu_khoa in DANH_SACH_TU_KHOA_DONG_XE:
        if _chuan_hoa_de_so_khop(tu_khoa) in t:
            return tu_khoa
    return None

def xe_thuoc_dong_xe_da_chon(ten_xe, ds_dong_xe):
    """Kiểm tra 1 xe (theo tên đầy đủ) có thuộc dòng xe (đúng 1 dòng xe cụ thể nhất)
    mà admin đã tick chọn hay không. Nhờ việc mỗi xe chỉ thuộc về 1 dòng xe duy nhất,
    khuyến mãi tick 'Air Blade' sẽ KHÔNG bị áp dụng nhầm sang xe 'Blade' và ngược lại."""
    if not ds_dong_xe:
        return False
    dx_cua_xe = _xac_dinh_dong_xe_cua_ten(ten_xe)
    return bool(dx_cua_xe) and dx_cua_xe in ds_dong_xe

_CACHE_KHUYEN_MAI = {'data': None, 'thoi_diem': 0}

def lay_danh_sach_khuyen_mai_dang_hoat_dong():
    """Lấy toàn bộ CTKM đang hoạt động (còn hạn + chưa bị tắt), cache ngắn hạn (5 giây)
    để tránh truy vấn CSDL liên tục khi trang chủ polling mỗi 3 giây."""
    now_ts = time.time()
    if _CACHE_KHUYEN_MAI['data'] is not None and (now_ts - _CACHE_KHUYEN_MAI['thoi_diem']) < 5:
        return _CACHE_KHUYEN_MAI['data']
    ds = [km for km in KhuyenMai.query.all() if km.dang_hoat_dong()]
    _CACHE_KHUYEN_MAI['data'] = ds
    _CACHE_KHUYEN_MAI['thoi_diem'] = now_ts
    return ds

def lay_khuyen_mai_ap_dung_cho_xe(ten_xe):
    """Trả về {'cty': [...], 'honda': [...]} gồm các CTKM đang hoạt động áp dụng cho xe này."""
    ket_qua = {'cty': [], 'honda': []}
    for km in lay_danh_sach_khuyen_mai_dang_hoat_dong():
        if xe_thuoc_dong_xe_da_chon(ten_xe, km.danh_sach_dong_xe):
            key = 'honda' if km.loai == 'honda' else 'cty'
            ket_qua[key].append(km.to_dict())
    return ket_qua

def seed_khu_vuc_giay_to_bl():
    """Tạo sẵn 3 khu vực lớn + các khu vực nhỏ Bạc Liêu nếu CSDL chưa có (idempotent)."""
    for idx_lon, kvl_data in enumerate(DU_LIEU_KHU_VUC_GIAY_TO_BL):
        kvl = KhuVucLonBL.query.filter_by(ma_khu_vuc=kvl_data['ma_khu_vuc']).first()
        if not kvl:
            kvl = KhuVucLonBL(
                ma_khu_vuc=kvl_data['ma_khu_vuc'],
                ten_khu_vuc=kvl_data['ten_khu_vuc'],
                thu_tu=idx_lon
            )
            db.session.add(kvl)
            db.session.flush()

        ten_nho_hien_co = {n.ten_khu_vuc_nho for n in KhuVucNhoBL.query.filter_by(khu_vuc_lon_id=kvl.id).all()}
        for idx_nho, ten_nho in enumerate(kvl_data['khu_vuc_nho']):
            if ten_nho not in ten_nho_hien_co:
                db.session.add(KhuVucNhoBL(khu_vuc_lon_id=kvl.id, ten_khu_vuc_nho=ten_nho, thu_tu=idx_nho))
    db.session.commit()

# --- HELPER FUNCTIONS & DECORATORS ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session or session.get('role') != 'admin':
            flash("Bạn không có quyền truy cập trang quản trị!", "danger")
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated_function

def save_image(file):
    """Lưu ảnh upload, đồng thời NÉN + THU NHỎ ảnh trước khi lưu.
    Đây là nguyên nhân chính khiến ảnh load chậm khi đã deploy lên server thật:
    ảnh chụp trực tiếp từ điện thoại thường rất nặng (3-8MB, độ phân giải 3000-4000px)
    trong khi hiển thị trên web chỉ cần vài trăm px, gây lãng phí băng thông rất lớn
    -> chậm rõ rệt trên mạng di động / server có băng thông hạn chế, dù chạy trên máy
    của mình (localhost) vẫn thấy nhanh vì không qua mạng thật.
    Nếu Pillow xử lý lỗi (file hỏng, định dạng lạ...), sẽ tự động lưu file gốc như cũ
    để không làm gián đoạn thao tác của admin."""
    if file and file.filename != '':
        filename = secure_filename(file.filename)
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        try:
            img = Image.open(file.stream)
            # Xoay ảnh đúng chiều theo thông tin EXIF (ảnh chụp điện thoại hay bị xoay
            # ngang nếu không xử lý bước này), đồng thời gỡ bỏ dữ liệu EXIF thừa (vị trí
            # GPS, thông tin thiết bị...) giúp giảm thêm dung lượng và bảo vệ quyền riêng tư.
            img = ImageOps.exif_transpose(img)

            # Giới hạn kích thước tối đa 1600px chiều dài nhất (đủ nét trên mọi màn hình
            # điện thoại/máy tính hiển thị danh sách xe/màu xe), không phóng to ảnh nhỏ hơn.
            MAX_KICH_THUOC = 1600
            img.thumbnail((MAX_KICH_THUOC, MAX_KICH_THUOC), Image.LANCZOS)

            ext = os.path.splitext(filename)[1].lower()
            if ext == '.png' and (img.mode in ('RGBA', 'LA') or 'transparency' in img.info):
                # Giữ định dạng PNG nếu ảnh có nền trong suốt, chỉ tối ưu nén (không mất chi tiết)
                img.save(save_path, format='PNG', optimize=True)
            else:
                # Các trường hợp còn lại nén sang JPEG chất lượng cao (82%) - giảm dung lượng
                # rất nhiều (thường còn 10-20% so với ảnh gốc) mà mắt thường không thấy khác biệt
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                img.save(save_path, format='JPEG', quality=82, optimize=True, progressive=True)
        except Exception as e:
            print("Lỗi nén ảnh, lưu file gốc:", e)
            file.stream.seek(0)
            file.save(save_path)
        return filename
    return ''

def lay_danh_sach_ma_mau(ten_mau):
    if not ten_mau: 
        return ['#cccccc']
    
    ten_mau_norm = unicodedata.normalize('NFC', str(ten_mau))
    text_val = re.sub(r'\([^)]*\)', '', ten_mau_norm).lower().strip()
    
    bang_mau = {
        'đỏ': '#ff0000', 'do': '#ff0000',
        'đen': '#000000', 'den': '#000000', 'đen nhám': '#222222', 'den nham': '#222222', 'nhám': '#222222', 'nham': '#222222',
        'trắng': '#ffffff', 'trang': '#ffffff', 'trắng ngọc': '#f8f9fa', 'trang ngoc': '#f8f9fa', 'ngọc': '#f8f9fa', 'ngoc': '#f8f9fa',
        'xanh': '#0000ff', 'xanh dương': '#0056b3', 'xanh duong': '#0056b3', 'xanh đậm': '#001f3f', 'xanh dam': '#001f3f', 'đậm': '#001f3f', 'dam': '#001f3f',
        'bạc': '#c0c0c0', 'bac': '#c0c0c0', 'xám': '#808080', 'xam': '#808080', 'xám xi măng': '#6c757d', 'xam xi mang': '#6c757d', 'xi': '#6c757d', 'măng': '#6c757d',
        'vàng': '#ffc107', 'vang': '#ffc107', 'cam': '#fd7e14', 'hồng': '#e83e8c', 'hong': '#e83e8c', 'xám mờ': '#555555'
    }
    
    danh_sach_kq = []
    for tu_khoa, ma in sorted(bang_mau.items(), key=lambda x: len(x[0]), reverse=True):
        if tu_khoa in text_val:
            if ma not in danh_sach_kq: 
                danh_sach_kq.append(ma)
            text_val = text_val.replace(tu_khoa, ' ')
            
    tu_list = text_val.split()
    for tu in tu_list:
        tu_sach = tu.strip()
        if tu_sach in bang_mau:
            ma_hex = bang_mau[tu_sach]
            if ma_hex not in danh_sach_kq: 
                danh_sach_kq.append(ma_hex)
        elif tu.startswith('#'):
            if tu not in danh_sach_kq: 
                danh_sach_kq.append(tu)
                
    return danh_sach_kq if danh_sach_kq else ['#cccccc']

def safe_float(val, default=0.0):
    try:
        if pd.isna(val): return default
        if isinstance(val, (int, float)): return float(val)
        s = str(val).replace('.', '').replace(',', '').strip()
        return float(s) if s else default
    except Exception:
        return default

def safe_int(val, default=0):
    try:
        if pd.isna(val): return default
        if isinstance(val, (int, float)): return int(val)
        s = str(val).replace('.', '').replace(',', '').strip()
        return int(float(s)) if s else default
    except Exception:
        return default

def get_order_priority():
    return case(
        (Xe.loai_xe == 'Xe số', 1),
        (Xe.loai_xe == 'Xe ga', 2),
        (Xe.loai_xe.ilike('%côn tay%'), 3),
        else_=4
    )

def cap_nhat_thoi_gian_dong_bo(vung='Cà Mau', username=None):
    vn_time = datetime.now(timezone(timedelta(hours=7)))
    time_str = vn_time.strftime("%H:%M' %d/%m/%Y")
    
    is_bl = 'bạc liêu' in (vung or '').lower()
    key_time = 'last_updated_bl' if is_bl else 'last_updated_cm'
    key_user = 'last_user_bl' if is_bl else 'last_user_cm'
    
    setting_time = Setting.query.filter_by(key=key_time).first()
    if not setting_time:
        db.session.add(Setting(key=key_time, value=time_str))
    else:
        setting_time.value = time_str
        
    curr_user = username
    if not curr_user and has_request_context() and 'username' in session:
        curr_user = session.get('username')
        
    ten_hien_thi = "Hệ thống"
    if curr_user:
        user_obj = User.query.filter_by(username=curr_user).first()
        if user_obj and user_obj.ho_ten:
            ten_hien_thi = user_obj.ho_ten
        else:
            ten_hien_thi = curr_user
            
    setting_user = Setting.query.filter_by(key=key_user).first()
    if not setting_user:
        db.session.add(Setting(key=key_user, value=ten_hien_thi))
    else:
        setting_user.value = ten_hien_thi
            
    db.session.commit()

# --- ROUTES AUTH & USER ---
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username_input = request.form.get("username")
        password_input = request.form.get("password")
        user = User.query.filter_by(username=username_input).first()
        
        if user and check_password_hash(user.password, password_input):
            # Kiểm tra trạng thái duyệt tài khoản
            trang_thai_tk = getattr(user, 'trang_thai', 'approved')
            if trang_thai_tk == 'pending':
                flash("Tài khoản của bạn đang chờ quản trị viên duyệt", "danger")
                return render_template("login.html")
            elif trang_thai_tk == 'rejected':
                flash("Tài khoản của bạn đã bị từ chối truy cập hệ thống.", "danger")
                return render_template("login.html")

            session.clear()
            session.permanent = True 
            
            session['is_logged_in'] = True  
            session['username'] = user.username
            session['role'] = str(user.role or '').strip().lower()  
            session['vung'] = user.khu_vuc or 'Cà Mau'
            session['ho_ten'] = user.ho_ten or user.username
            
            return redirect(url_for('home'))
        flash("Sai thông tin đăng nhập!", "danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route("/set-vung", methods=["POST"])
def set_vung():
    if 'username' not in session:
        return redirect(url_for('login'))
    vung_moi = request.form.get("vung", "").strip()
    if vung_moi:
        session['vung'] = vung_moi
    return redirect(url_for('home'))

@app.route("/api/change-password", methods=["POST"])
def change_password():
    if 'username' not in session:
        return jsonify({"success": False, "message": "Phiên đăng nhập đã hết hạn, vui lòng đăng nhập lại."}), 401

    data = request.get_json(silent=True) or {}
    mat_khau_cu = (data.get('mat_khau_cu') or '').strip()
    mat_khau_moi = (data.get('mat_khau_moi') or '').strip()
    xac_nhan = (data.get('xac_nhan') or '').strip()

    if not mat_khau_cu or not mat_khau_moi or not xac_nhan:
        return jsonify({"success": False, "message": "Vui lòng nhập đầy đủ thông tin."}), 400

    if mat_khau_moi != xac_nhan:
        return jsonify({"success": False, "message": "Mật khẩu mới và xác nhận không khớp."}), 400

    if len(mat_khau_moi) < 6:
        return jsonify({"success": False, "message": "Mật khẩu mới phải có ít nhất 6 ký tự."}), 400

    user = User.query.filter_by(username=session['username']).first()
    if not user:
        return jsonify({"success": False, "message": "Không tìm thấy tài khoản."}), 404

    if not check_password_hash(user.password, mat_khau_cu):
        return jsonify({"success": False, "message": "Mật khẩu hiện tại không đúng."}), 400

    if check_password_hash(user.password, mat_khau_moi):
        return jsonify({"success": False, "message": "Mật khẩu mới phải khác mật khẩu hiện tại."}), 400

    user.password = generate_password_hash(mat_khau_moi)
    db.session.commit()

    return jsonify({"success": True, "message": "Đổi mật khẩu thành công!"})

@app.route("/home")
def home():
    if 'username' not in session: 
        return redirect(url_for('login'))
    
    current_user = User.query.filter_by(username=session['username']).first()
    khu_vuc_user = (session.get('vung') or (current_user.khu_vuc if current_user else 'Cà Mau')).strip()
    
    is_bl = 'bạc liêu' in khu_vuc_user.lower()
    key_time = 'last_updated_bl' if is_bl else 'last_updated_cm'
    key_user = 'last_user_bl' if is_bl else 'last_user_cm'
    
    setting_time = Setting.query.filter_by(key=key_time).first()
    setting_user = Setting.query.filter_by(key=key_user).first()
    
    last_updated_str = setting_time.value if (setting_time and setting_time.value) else "Chưa cập nhật"
    last_user_str = setting_user.value if (setting_user and setting_user.value) else "Hệ thống"
    
    search_query = request.args.get('search', '')
    loai_filter = request.args.get('loai', '')
    
    query = Xe.query
    if search_query:
        query = query.filter((Xe.ten_xe.ilike(f'%{search_query}%')) | (Xe.phien_ban.ilike(f'%{search_query}%')))
    if loai_filter:
        query = query.filter_by(loai_xe=loai_filter)
        
    danh_sach_xe = query.order_by(get_order_priority(), Xe.ten_xe.asc()).all()
    danh_sach_loai = [l[0] for l in db.session.query(Xe.loai_xe).distinct().all() if l[0]]
    data = [format_xe_data_home(xe, khu_vuc_user) for xe in danh_sach_xe]
       
    return render_template(
        "home.html", 
        danh_sach_xe=data, 
        search_query=search_query, 
        loai_filter=loai_filter, 
        danh_sach_loai=danh_sach_loai, 
        username=session.get('username'),
        current_user=current_user,
        user_vung=khu_vuc_user,
        last_updated=last_updated_str,
        last_updated_by=last_user_str
    )

# --- ADMIN ROUTES ---
@app.route("/admin")
@admin_required
def admin_panel():
    search_query = request.args.get('search', '')
    loai_filter = request.args.get('loai', '')
    
    query = Xe.query
    if search_query:
        query = query.filter((Xe.ten_xe.ilike(f'%{search_query}%')) | (Xe.phien_ban.ilike(f'%{search_query}%')))
    if loai_filter:
        query = query.filter_by(loai_xe=loai_filter)
        
    danh_sach_xe = query.order_by(get_order_priority(), Xe.ten_xe.asc()).all()
    danh_sach_loai = [l[0] for l in db.session.query(Xe.loai_xe).distinct().all() if l[0]]
    
    setting = Setting.query.filter_by(key='csv_url').first()
    csv_url = setting.value if setting else ''

    # --- CHỈ LẤY CÁC TÀI KHOẢN ĐANG CHỜ DUYỆT (trang_thai = 'pending') ---
    danh_sach_user = User.query.filter_by(trang_thai='pending').order_by(User.id.asc()).all()

    # --- DANH SÁCH CHƯƠNG TRÌNH KHUYẾN MÃI (mới nhất lên trước) ---
    danh_sach_khuyen_mai = KhuyenMai.query.order_by(KhuyenMai.ngay_bat_dau.desc(), KhuyenMai.id.desc()).all()
    danh_sach_dong_xe = lay_danh_sach_dong_xe_hien_co()

    return render_template(
        "admin.html", 
        danh_sach_xe=danh_sach_xe, 
        danh_sach_user=danh_sach_user,  # <--- Truyền biến này ra giao diện admin.html
        danh_sach_khuyen_mai=danh_sach_khuyen_mai,
        danh_sach_dong_xe=danh_sach_dong_xe,
        search_query=search_query, 
        loai_filter=loai_filter, 
        danh_sach_loai=danh_sach_loai, 
        username=session.get('username'),
        csv_url=csv_url
    )

# --- QUẢN LÝ CHƯƠNG TRÌNH KHUYẾN MÃI ---
def _doc_ngay(s):
    s = (s or '').strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None

@app.route("/admin/khuyen-mai/add", methods=["POST"])
@admin_required
def them_khuyen_mai():
    tieu_de = request.form.get("tieu_de", "").strip()
    loai = request.form.get("loai", "cty").strip()
    if loai not in ('cty', 'honda'):
        loai = 'cty'
    ngay_bat_dau = _doc_ngay(request.form.get("ngay_bat_dau"))
    ngay_ket_thuc = _doc_ngay(request.form.get("ngay_ket_thuc"))
    ds_dong_xe = request.form.getlist("dong_xe[]")
    noi_dung = request.form.get("noi_dung", "").strip()

    if not tieu_de or not ngay_bat_dau or not ngay_ket_thuc:
        flash("Vui lòng nhập đầy đủ Tiêu đề, Ngày bắt đầu và Ngày kết thúc!", "danger")
        return redirect(url_for('admin_panel'))
    if ngay_ket_thuc < ngay_bat_dau:
        flash("Ngày kết thúc phải sau hoặc bằng Ngày bắt đầu!", "danger")
        return redirect(url_for('admin_panel'))
    if not ds_dong_xe:
        flash("Vui lòng tick chọn ít nhất 1 dòng xe được áp dụng!", "danger")
        return redirect(url_for('admin_panel'))

    km = KhuyenMai(
        loai=loai,
        tieu_de=tieu_de,
        noi_dung=noi_dung,
        ngay_bat_dau=ngay_bat_dau,
        ngay_ket_thuc=ngay_ket_thuc,
        dang_bat=True,
        created_by=session.get('username')
    )
    km.danh_sach_dong_xe = ds_dong_xe
    db.session.add(km)
    db.session.commit()
    flash(f"Đã thêm chương trình khuyến mãi '{tieu_de}'!", "success")
    return redirect(url_for('admin_panel'))

@app.route("/admin/khuyen-mai/edit/<int:id>", methods=["POST"])
@admin_required
def sua_khuyen_mai(id):
    km = db.get_or_404(KhuyenMai, id)
    tieu_de = request.form.get("tieu_de", "").strip()
    loai = request.form.get("loai", km.loai).strip()
    if loai not in ('cty', 'honda'):
        loai = km.loai
    ngay_bat_dau = _doc_ngay(request.form.get("ngay_bat_dau"))
    ngay_ket_thuc = _doc_ngay(request.form.get("ngay_ket_thuc"))
    ds_dong_xe = request.form.getlist("dong_xe[]")
    noi_dung = request.form.get("noi_dung", "").strip()

    if not tieu_de or not ngay_bat_dau or not ngay_ket_thuc:
        flash("Vui lòng nhập đầy đủ Tiêu đề, Ngày bắt đầu và Ngày kết thúc!", "danger")
        return redirect(url_for('admin_panel'))
    if ngay_ket_thuc < ngay_bat_dau:
        flash("Ngày kết thúc phải sau hoặc bằng Ngày bắt đầu!", "danger")
        return redirect(url_for('admin_panel'))
    if not ds_dong_xe:
        flash("Vui lòng tick chọn ít nhất 1 dòng xe được áp dụng!", "danger")
        return redirect(url_for('admin_panel'))

    km.tieu_de = tieu_de
    km.loai = loai
    km.noi_dung = noi_dung
    km.ngay_bat_dau = ngay_bat_dau
    km.ngay_ket_thuc = ngay_ket_thuc
    km.danh_sach_dong_xe = ds_dong_xe
    db.session.commit()
    flash(f"Đã cập nhật chương trình khuyến mãi '{tieu_de}'!", "success")
    return redirect(url_for('admin_panel'))

@app.route("/admin/khuyen-mai/toggle/<int:id>", methods=["GET"])
@admin_required
def bat_tat_khuyen_mai(id):
    km = db.get_or_404(KhuyenMai, id)
    km.dang_bat = not km.dang_bat
    db.session.commit()
    flash(f"Đã {'bật' if km.dang_bat else 'tắt'} chương trình khuyến mãi '{km.tieu_de}'!", "success")
    return redirect(url_for('admin_panel'))

@app.route("/admin/khuyen-mai/delete/<int:id>", methods=["GET"])
@admin_required
def xoa_khuyen_mai(id):
    km = db.get_or_404(KhuyenMai, id)
    ten = km.tieu_de
    db.session.delete(km)
    db.session.commit()
    flash(f"Đã xóa chương trình khuyến mãi '{ten}'!", "success")
    return redirect(url_for('admin_panel'))

@app.context_processor
def inject_update_info():
    t_cm = Setting.query.filter_by(key='last_updated_cm').first()
    u_cm = Setting.query.filter_by(key='last_user_cm').first()
    
    t_bl = Setting.query.filter_by(key='last_updated_bl').first()
    u_bl = Setting.query.filter_by(key='last_user_bl').first()

    return {
        'thoi_gian_camau': t_cm.value if t_cm else 'Chưa cập nhật',
        'nguoi_camau': u_cm.value if u_cm else 'Admin',
        'thoi_gian_baclieu': t_bl.value if t_bl else 'Chưa cập nhật',
        'nguoi_baclieu': u_bl.value if u_bl else 'Admin'
    }

@app.route("/admin/settings", methods=["POST"])
@admin_required

def save_settings():
    csv_url = request.form.get("csv_url", "").strip()
    setting = Setting.query.filter_by(key='csv_url').first()
    if not setting:
        setting = Setting(key='csv_url', value=csv_url)
        db.session.add(setting)
    else:
        setting.value = csv_url
    db.session.commit()
    flash("Lưu link Google Sheets thành công!", "success")
    return redirect(url_for('admin_panel'))

def _chuan_hoa_ten_xe_bo_qua_moi(ten_xe):
    """
    Chuẩn hoá tên xe để so khớp 2 xe THỰC CHẤT LÀ 1, chỉ khác việc sheet có/không
    có hậu tố "(MỚI)" ở cuối theo từng tháng (VD: "... ACB125K2VM" và "... ACB125K2VM (MỚI)").
    """
    s = unicodedata.normalize('NFC', str(ten_xe or '')).strip()
    # Bắt cả "(MỚI)" có ngoặc lẫn "MỚI" không ngoặc ở cuối tên (VD: "...K2CN MỚI")
    s = re.sub(r'\s*\(?\s*m[ớo]i\s*\)?\s*$', '', s, flags=re.IGNORECASE)
    return s.strip().lower()

def _gop_xe_trung_ten_bo_qua_moi():
    """
    Quét toàn bộ xe hiện có trong CSDL, gộp các xe có tên giống hệt nhau sau khi bỏ hậu tố "(MỚI)"
    (VD: cùng là 1 xe nhưng tháng trước sheet đặt tên có "(MỚI)", tháng này bỏ đi).
    Xe không có hậu tố "(MỚI)" được coi là bản ghi CHÍNH (tên hiện hành) và được giữ lại;
    mọi giá xe / giá giấy tờ / màu xe của bản ghi phụ sẽ được chuyển sang bản ghi chính
    (chỉ điền vào những chỗ bản ghi chính đang trống/0, không ghi đè giá trị đã có),
    sau đó xoá bản ghi phụ. Nhờ vậy giá giấy tờ đã nhập không bị mất khi tên xe đổi qua các tháng.
    Trả về số cặp đã gộp.
    """
    nhom_theo_ten = {}
    for xe in Xe.query.all():
        key = _chuan_hoa_ten_xe_bo_qua_moi(xe.ten_xe)
        nhom_theo_ten.setdefault(key, []).append(xe)

    so_luong_gop = 0
    for key, ds_xe in nhom_theo_ten.items():
        if len(ds_xe) < 2:
            continue

        # Ưu tiên giữ lại xe có tên KHÔNG mang hậu tố "(MỚI)"/"MỚI" làm bản ghi chính
        xe_khong_hau_to_moi = [
            x for x in ds_xe
            if not re.search(r'\(?\s*m[ớo]i\s*\)?\s*$', x.ten_xe.strip(), re.IGNORECASE)
        ]
        xe_chinh = xe_khong_hau_to_moi[0] if xe_khong_hau_to_moi else ds_xe[0]
        ds_xe_phu = [x for x in ds_xe if x.id != xe_chinh.id]

        for xe_phu in ds_xe_phu:
            # Giá xe (chỉ lấy khi bản ghi chính đang trống/0)
            if not xe_chinh.gia_cm_cao: xe_chinh.gia_cm_cao = xe_phu.gia_cm_cao
            if not xe_chinh.gia_cm_trung: xe_chinh.gia_cm_trung = xe_phu.gia_cm_trung
            if not xe_chinh.gia_cm_thap: xe_chinh.gia_cm_thap = xe_phu.gia_cm_thap
            if not xe_chinh.gia_bl_cao: xe_chinh.gia_bl_cao = xe_phu.gia_bl_cao
            if not xe_chinh.gia_bl_trung: xe_chinh.gia_bl_trung = xe_phu.gia_bl_trung
            if not xe_chinh.gia_bl_thap: xe_chinh.gia_bl_thap = xe_phu.gia_bl_thap

            # Giá giấy tờ Cà Mau
            if not xe_chinh.gia_gt_phuong_cm: xe_chinh.gia_gt_phuong_cm = xe_phu.gia_gt_phuong_cm
            if not xe_chinh.gia_gt_xa_cm: xe_chinh.gia_gt_xa_cm = xe_phu.gia_gt_xa_cm
            if not xe_chinh.gia_gt_phuong_bl: xe_chinh.gia_gt_phuong_bl = xe_phu.gia_gt_phuong_bl
            if not xe_chinh.gia_gt_xa_bl: xe_chinh.gia_gt_xa_bl = xe_phu.gia_gt_xa_bl

            # Giá giấy tờ Bạc Liêu theo từng khu vực nhỏ: chuyển toàn bộ sang xe chính
            for g in GiaGiayToXeBL.query.filter_by(xe_id=xe_phu.id).all():
                g_chinh = GiaGiayToXeBL.query.filter_by(
                    xe_id=xe_chinh.id, khu_vuc_nho_id=g.khu_vuc_nho_id
                ).first()
                if g_chinh:
                    if not g_chinh.gia:
                        g_chinh.gia = g.gia
                    db.session.delete(g)
                else:
                    g.xe_id = xe_chinh.id

            # Màu xe: giữ màu của xe chính nếu trùng tên màu, ngược lại chuyển sang xe chính
            for mau in XeMau.query.filter_by(xe_id=xe_phu.id).all():
                mau_chinh = XeMau.query.filter_by(xe_id=xe_chinh.id, ten_mau=mau.ten_mau).first()
                if mau_chinh:
                    db.session.delete(mau)
                else:
                    mau.xe_id = xe_chinh.id

            db.session.delete(xe_phu)
            so_luong_gop += 1

    if so_luong_gop:
        db.session.commit()

    return so_luong_gop

_TU_KHOA_PHAN_LOAI_VISION = ['thể thao', 'cao cấp', 'đặc biệt', 'tiêu chuẩn']

def _phan_loai_vision(ten_xe):
    """Xác định phân loại VISION (Tiêu chuẩn/Đặc biệt/Cao cấp/Thể thao) dựa trên tên xe."""
    t = unicodedata.normalize('NFC', str(ten_xe or '')).lower()
    for tu_khoa in _TU_KHOA_PHAN_LOAI_VISION:
        if tu_khoa in t:
            return tu_khoa
    return None

def _don_dep_xe_vision_chi_tiet(ten_vision_da_gop):
    """
    Dọn các xe VISION "chi tiết" (theo từng miền Bắc/Nam riêng, tên đặt kiểu
    "V06-[V01]", "V01-V06]"... khác hẳn tên đã cộng gộp) còn sót lại trong CSDL từ những
    lần đồng bộ trước khi có bộ lọc theo dòng "Cộng gộp". Mỗi xe chi tiết được nhận diện
    PHÂN LOẠI (Tiêu chuẩn/Đặc biệt/Cao cấp/Thể thao) rồi gộp dữ liệu (giá xe, giá giấy tờ,
    màu xe - chỉ điền vào chỗ đang trống/0) vào đúng bản ĐÃ CỘNG GỘP cùng phân loại của
    tháng này, sau đó xoá bản chi tiết. Xe không xác định được phân loại hoặc không có bản
    cộng gộp tương ứng sẽ được GIỮ NGUYÊN (không xoá nhầm).
    Trả về số xe chi tiết đã dọn.
    """
    if not ten_vision_da_gop:
        return 0

    tat_ca_xe_vision = Xe.query.filter(Xe.ten_xe.ilike('%vision%')).all()

    xe_chinh_theo_loai = {}
    for xe in tat_ca_xe_vision:
        if xe.ten_xe in ten_vision_da_gop:
            loai = _phan_loai_vision(xe.ten_xe)
            if loai and loai not in xe_chinh_theo_loai:
                xe_chinh_theo_loai[loai] = xe

    so_luong_don = 0
    for xe in tat_ca_xe_vision:
        if xe.ten_xe in ten_vision_da_gop:
            continue  # đây là bản đã cộng gộp của tháng này -> giữ nguyên

        loai = _phan_loai_vision(xe.ten_xe)
        xe_chinh = xe_chinh_theo_loai.get(loai)
        if not xe_chinh:
            continue  # không xác định được phân loại / chưa có bản cộng gộp tương ứng -> bỏ qua

        if not xe_chinh.gia_cm_cao: xe_chinh.gia_cm_cao = xe.gia_cm_cao
        if not xe_chinh.gia_cm_trung: xe_chinh.gia_cm_trung = xe.gia_cm_trung
        if not xe_chinh.gia_cm_thap: xe_chinh.gia_cm_thap = xe.gia_cm_thap
        if not xe_chinh.gia_bl_cao: xe_chinh.gia_bl_cao = xe.gia_bl_cao
        if not xe_chinh.gia_bl_trung: xe_chinh.gia_bl_trung = xe.gia_bl_trung
        if not xe_chinh.gia_bl_thap: xe_chinh.gia_bl_thap = xe.gia_bl_thap
        if not xe_chinh.gia_gt_phuong_cm: xe_chinh.gia_gt_phuong_cm = xe.gia_gt_phuong_cm
        if not xe_chinh.gia_gt_xa_cm: xe_chinh.gia_gt_xa_cm = xe.gia_gt_xa_cm
        if not xe_chinh.gia_gt_phuong_bl: xe_chinh.gia_gt_phuong_bl = xe.gia_gt_phuong_bl
        if not xe_chinh.gia_gt_xa_bl: xe_chinh.gia_gt_xa_bl = xe.gia_gt_xa_bl

        for g in GiaGiayToXeBL.query.filter_by(xe_id=xe.id).all():
            g_chinh = GiaGiayToXeBL.query.filter_by(
                xe_id=xe_chinh.id, khu_vuc_nho_id=g.khu_vuc_nho_id
            ).first()
            if g_chinh:
                if not g_chinh.gia:
                    g_chinh.gia = g.gia
                db.session.delete(g)
            else:
                g.xe_id = xe_chinh.id

        for mau in XeMau.query.filter_by(xe_id=xe.id).all():
            mau_chinh = XeMau.query.filter_by(xe_id=xe_chinh.id, ten_mau=mau.ten_mau).first()
            if mau_chinh:
                db.session.delete(mau)
            else:
                mau.xe_id = xe_chinh.id

        db.session.delete(xe)
        so_luong_don += 1

    if so_luong_don:
        db.session.commit()

    return so_luong_don

_sync_lock = threading.Lock()

def run_sync_process(username=None):
    # Không cho 2 lần đồng bộ (nền + tay) chạy chồng lên nhau cùng lúc,
    # tránh tạo trùng xe / ghi đè tồn kho không nhất quán.
    if not _sync_lock.acquire(blocking=False):
        return False, "Đang có một tiến trình đồng bộ khác chạy, vui lòng thử lại sau."
    try:
        return _run_sync_process_inner(username=username)
    finally:
        _sync_lock.release()

def _run_sync_process_inner(username=None):
    setting = Setting.query.filter_by(key='csv_url').first()
    csv_url = setting.value if setting else None
    
    if not csv_url:
        return False, "Chưa cấu hình đường dẫn Google Sheets CSV!"
        
    try:
        if 'pubhtml' in csv_url:
            csv_url = csv_url.replace('/pubhtml', '/pub')
        if 'output=csv' not in csv_url:
            separator = '&' if '?' in csv_url else '?'
            csv_url += f'{separator}output=csv'
        
        cache_buster = f"&_t={int(time.time())}"
        csv_url += cache_buster

        try:
            df_raw = pd.read_csv(csv_url, header=None, on_bad_lines='skip')
        except Exception as e:
            return False, f"Không thể tải dữ liệu từ URL: {str(e)}"

        valid_vehicle_keywords = [
            'wave', 'blade', 'future', 'vision', 'air blade', 'sh', 'winner', 
            'lead', 'vario', 'cub', 'rebel', 'cb', 'cbr', 'afb', 'afs', 'supream'
        ]

        # Gộp các xe trùng nhau chỉ khác hậu tố "(MỚI)" trước khi đối chiếu với sheet mới,
        # để tên xe đổi qua từng tháng không tạo ra bản ghi (và phải nhập lại giá giấy tờ) mới.
        so_luong_gop_ten = _gop_xe_trung_ten_bo_qua_moi()

        all_xe_db = Xe.query.all()
        all_xe_dict = {xe.ten_xe: xe for xe in all_xe_db}
        
        garbage_xe_ids = [
            xe.id for xe in all_xe_db 
            if not any(kw in xe.ten_xe.lower() for kw in valid_vehicle_keywords)
        ]
        if garbage_xe_ids:
            XeMau.query.filter(XeMau.xe_id.in_(garbage_xe_ids)).delete(synchronize_session=False)
            Xe.query.filter(Xe.id.in_(garbage_xe_ids)).delete(synchronize_session=False)
            db.session.commit()
            all_xe_dict = {xe.ten_xe: xe for xe in Xe.query.all()}

        all_mau_dict = {(m.xe_id, m.ten_mau): m for m in XeMau.query.all()}

        header_row_idx = None
        for i, row in df_raw.iterrows():
            row_str = " ".join([str(val).upper() for val in row.values])
            if ('DÒNG XE' in row_str or 'TÊN XE' in row_str) and 'MÀU' in row_str:
                header_row_idx = i
                break
        
        df = pd.read_csv(csv_url, header=header_row_idx if header_row_idx is not None else 1, on_bad_lines='skip')
        df = df.dropna(how='all')
        cols = [str(c).strip().upper() for c in df.columns]
        
        col_xe_idx, col_mau_idx = 0, 1
        for idx, col_name in enumerate(cols):
            clean_name = col_name.split('.')[0].strip()
            if 'DÒNG XE' in clean_name or 'TÊN XE' in clean_name:
                col_xe_idx = idx
            elif clean_name == 'MÀU':
                col_mau_idx = idx

        col_ns1_idx = col_mau_idx + 1
        col_ns2_idx = col_mau_idx + 2
        col_ns3_idx = col_mau_idx + 3
        col_ns4_idx = col_mau_idx + 4
        col_ns5_idx = col_mau_idx + 5
        col_nsm1_idx = col_mau_idx + 6

        col_gia_cm_cao_idx = -1
        col_gia_bl_cao_idx = -1
        for idx, col_name in enumerate(cols):
            if 'gia_cm_cao' in col_name.lower(): col_gia_cm_cao_idx = idx
            if 'gia_bl_cao' in col_name.lower(): col_gia_bl_cao_idx = idx

        if col_xe_idx < df.shape[1]:
            df.iloc[:, col_xe_idx] = df.iloc[:, col_xe_idx].ffill()

        def parse_stock(val):
            if pd.isna(val): return 0
            try:
                s = str(val).strip()
                if not s or s.lower() in ['nan', 'none', '']: return 0
                num = int(float(s.replace(',', '').replace('.0', '')))
                return num if 0 <= num <= 99 else 0
            except:
                return 0

        def _co_cum_cong_gop_vision(text):
            t = unicodedata.normalize('NFC', str(text or '')).lower()
            return 'cộng gộp' in t and 'vision' in t

        # Tìm dòng đánh dấu "Cộng gộp các dòng Vision của miền Bắc và Miền Nam..." trong sheet.
        # Các dòng VISION nằm TRƯỚC dòng đánh dấu này là dữ liệu CHI TIẾT theo từng miền (Bắc/Nam
        # riêng lẻ) -> bỏ qua, không tạo thành xe. Chỉ lấy các dòng VISION nằm TỪ dòng đánh dấu
        # trở về sau (đã được cộng gộp Bắc+Nam) để tránh tạo trùng nhiều "xe" cùng phiên bản/giá.
        vi_tri_dong_gop_vision = None
        for idx in range(len(df)):
            dong_gop_lai = " ".join([str(v) for v in df.iloc[idx].values])
            if _co_cum_cong_gop_vision(dong_gop_lai):
                vi_tri_dong_gop_vision = idx
                break

        processed_data = []
        for idx in range(len(df)):
            if col_xe_idx >= df.shape[1] or col_mau_idx >= df.shape[1]:
                continue
                
            ten_xe_excel = str(df.iloc[idx, col_xe_idx]).strip()
            if not ten_xe_excel or ten_xe_excel.lower() in ['nan', 'none', 'tổng', 'total', 'unnamed', '0']: 
                continue
            if len(ten_xe_excel) > 150 or not any(kw in ten_xe_excel.lower() for kw in valid_vehicle_keywords):
                continue
            if _co_cum_cong_gop_vision(ten_xe_excel):
                continue

            # Bỏ qua các dòng VISION chi tiết (chưa cộng gộp Bắc/Nam) đứng trước dòng đánh dấu
            if (vi_tri_dong_gop_vision is not None and idx < vi_tri_dong_gop_vision
                    and 'vision' in ten_xe_excel.lower()):
                continue

            ten_mau_excel = str(df.iloc[idx, col_mau_idx]).strip()
            if not ten_mau_excel or ten_mau_excel.lower() in ['nan', 'none', 'unnamed', '0']:
                continue
            
            ns1 = parse_stock(df.iloc[idx, col_ns1_idx]) if col_ns1_idx < df.shape[1] else 0
            ns2 = parse_stock(df.iloc[idx, col_ns2_idx]) if col_ns2_idx < df.shape[1] else 0
            ns3 = parse_stock(df.iloc[idx, col_ns3_idx]) if col_ns3_idx < df.shape[1] else 0
            ns4 = parse_stock(df.iloc[idx, col_ns4_idx]) if col_ns4_idx < df.shape[1] else 0
            ns5 = parse_stock(df.iloc[idx, col_ns5_idx]) if col_ns5_idx < df.shape[1] else 0
            nsm1 = parse_stock(df.iloc[idx, col_nsm1_idx]) if col_nsm1_idx < df.shape[1] else 0

            gia_cm_moi = safe_float(df.iloc[idx, col_gia_cm_cao_idx]) if col_gia_cm_cao_idx != -1 and col_gia_cm_cao_idx < df.shape[1] else None
            gia_bl_moi = safe_float(df.iloc[idx, col_gia_bl_cao_idx]) if col_gia_bl_cao_idx != -1 and col_gia_bl_cao_idx < df.shape[1] else None

            processed_data.append({
                'ten_xe': ten_xe_excel,
                'ten_mau': ten_mau_excel,
                'ns1': ns1, 'ns2': ns2, 'ns3': ns3, 'ns4': ns4, 'ns5': ns5, 'nsm1': nsm1,
                'gia_cm_moi': gia_cm_moi,
                'gia_bl_moi': gia_bl_moi
            })

        if not processed_data:
            return False, "Không tìm thấy dữ liệu hợp lệ trong Google Sheets."

        # Dọn các xe VISION "chi tiết" (theo miền, tên khác kiểu "V06-[V01]"...) còn sót lại
        # trong CSDL từ trước, gộp dữ liệu về đúng bản đã cộng gộp (chỉ có 4 phiên bản/tháng)
        ten_vision_da_gop = {d['ten_xe'] for d in processed_data if 'vision' in d['ten_xe'].lower()}
        so_luong_don_vision = _don_dep_xe_vision_chi_tiet(ten_vision_da_gop)
        if so_luong_don_vision:
            all_xe_db = Xe.query.all()
            all_xe_dict = {xe.ten_xe: xe for xe in all_xe_db}

        df_clean = pd.DataFrame(processed_data)
        
        so_luong_them = 0
        so_luong_cap_nhat = 0
        xe_inventory_map = {}
        new_xe_objects = []
        
        has_price_changed_cm = False
        has_price_changed_bl = False

        # Map tên đã chuẩn hoá (bỏ hậu tố "(MỚI)") -> danh sách xe, dùng để nhận diện xe
        # trong sheet mới có tên chỉ khác xe đã có trong CSDL đúng mỗi hậu tố "(MỚI)"
        all_xe_dict_norm = {}
        for ten, xe_obj in all_xe_dict.items():
            key_norm = _chuan_hoa_ten_xe_bo_qua_moi(ten)
            all_xe_dict_norm.setdefault(key_norm, []).append(xe_obj)

        for _, row in df_clean.iterrows():
            ten_xe_excel = row['ten_xe']
            loai_xe_excel = tu_dong_phan_loai(ten_xe_excel)
            
            if ten_xe_excel not in all_xe_dict:
                # Thử tìm xe đã có trong CSDL nhưng tên chỉ khác đúng hậu tố "(MỚI)"
                # -> coi là CÙNG 1 xe, đổi tên theo sheet tháng này thay vì tạo bản ghi mới
                # (giữ nguyên giá xe / giá giấy tờ / màu xe đã nhập trước đó).
                key_norm = _chuan_hoa_ten_xe_bo_qua_moi(ten_xe_excel)
                ung_vien = [x for x in all_xe_dict_norm.get(key_norm, []) if x.ten_xe in all_xe_dict]
                if len(ung_vien) == 1:
                    xe = ung_vien[0]
                    ten_cu = xe.ten_xe
                    if ten_cu != ten_xe_excel:
                        del all_xe_dict[ten_cu]
                        xe.ten_xe = ten_xe_excel
                        all_xe_dict[ten_xe_excel] = xe
                    if loai_xe_excel and xe.loai_xe == "Chưa phân loại":
                        xe.loai_xe = loai_xe_excel
                    so_luong_cap_nhat += 1
                else:
                    xe = Xe(loai_xe=loai_xe_excel, ten_xe=ten_xe_excel, phien_ban='')
                    new_xe_objects.append(xe)
                    db.session.add(xe)
                    all_xe_dict[ten_xe_excel] = xe
                    so_luong_them += 1
            else:
                xe = all_xe_dict[ten_xe_excel]
                if loai_xe_excel and xe.loai_xe == "Chưa phân loại":
                    xe.loai_xe = loai_xe_excel
                so_luong_cap_nhat += 1

            if row.get('gia_cm_moi') is not None and row['gia_cm_moi'] > 0:
                if xe.gia_cm_cao != row['gia_cm_moi']:
                    xe.gia_cm_cao = row['gia_cm_moi']
                    has_price_changed_cm = True
            if row.get('gia_bl_moi') is not None and row['gia_bl_moi'] > 0:
                if xe.gia_bl_cao != row['gia_bl_moi']:
                    xe.gia_bl_cao = row['gia_bl_moi']
                    has_price_changed_bl = True

        if new_xe_objects:
            db.session.flush()

        all_mau_dict = {(m.xe_id, m.ten_mau): m for m in XeMau.query.all()}

        for _, row in df_clean.iterrows():
            xe = all_xe_dict[row['ten_xe']]
            ten_mau_excel = row['ten_mau']
            ns1, ns2, ns3 = int(row['ns1']), int(row['ns2']), int(row['ns3'])
            ns4, ns5, nsm1 = int(row['ns4']), int(row['ns5']), int(row['nsm1'])

            key_mau = (xe.id, ten_mau_excel)
            if key_mau not in all_mau_dict:
                mau_existing = XeMau(xe_id=xe.id, ten_mau=ten_mau_excel, ns1=ns1, ns2=ns2, ns3=ns3, ns4=ns4, ns5=ns5, nsm1=nsm1)
                db.session.add(mau_existing)
                all_mau_dict[key_mau] = mau_existing
            else:
                mau_existing = all_mau_dict[key_mau]
                mau_existing.ns1, mau_existing.ns2, mau_existing.ns3 = ns1, ns2, ns3
                mau_existing.ns4, mau_existing.ns5, mau_existing.nsm1 = ns4, ns5, nsm1
            
            if xe.id not in xe_inventory_map:
                xe_inventory_map[xe.id] = {'ns1': 0, 'ns2': 0, 'ns3': 0, 'ns4': 0, 'ns5': 0, 'nsm1': 0}
            
            xe_inventory_map[xe.id]['ns1'] += ns1
            xe_inventory_map[xe.id]['ns2'] += ns2
            xe_inventory_map[xe.id]['ns3'] += ns3
            xe_inventory_map[xe.id]['ns4'] += ns4
            xe_inventory_map[xe.id]['ns5'] += ns5
            xe_inventory_map[xe.id]['nsm1'] += nsm1

        for xe_id, inv in xe_inventory_map.items():
            xe_obj = db.session.get(Xe, xe_id)
            if xe_obj:
                xe_obj.ns1, xe_obj.ns2, xe_obj.ns3 = inv['ns1'], inv['ns2'], inv['ns3']
                xe_obj.ns4, xe_obj.ns5, xe_obj.nsm1 = inv['ns4'], inv['ns5'], inv['nsm1']

        db.session.commit()
        
        if has_price_changed_cm:
            cap_nhat_thoi_gian_dong_bo("Cà Mau", username)
        if has_price_changed_bl:
            cap_nhat_thoi_gian_dong_bo("Bạc Liêu", username)
            
        thong_bao_gop = f" (Đã gộp {so_luong_gop_ten} xe trùng tên do đổi hậu tố '(MỚI)'.)" if so_luong_gop_ten else ""
        thong_bao_vision = f" (Đã dọn {so_luong_don_vision} xe VISION chi tiết, gộp về bản cộng gộp.)" if so_luong_don_vision else ""
        return True, f"Thêm mới {so_luong_them} xe, Cập nhật {so_luong_cap_nhat} xe.{thong_bao_gop}{thong_bao_vision}"
        
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return False, str(e)

# Chu kỳ đồng bộ nền bình thường (giây). 6 giây/lần là quá dày, dễ bị Google
# Sheets giới hạn (429) và tốn tài nguyên DB không cần thiết -> giãn ra 2 phút.
# SYNC_INTERVAL_SECONDS = 120
# # Khi lỗi liên tiếp (VD: link sai, mất mạng), lùi thời gian chờ ra để tránh
# # spam request/log liên tục.
# SYNC_ERROR_BACKOFF_SECONDS = 300

# def start_background_sync():
#     def run_loop():
#         time.sleep(10)  # đợi app khởi động xong trước lần sync đầu tiên
#         while True:
#             try:
#                 with app.app_context():
#                     success, msg = run_sync_process(username=None)
#                     print(f"--- [BACKGROUND SYNC] {msg} ---")
#                 time.sleep(SYNC_INTERVAL_SECONDS if success else SYNC_ERROR_BACKOFF_SECONDS)
#             except Exception as e:
#                 print(f"[BACKGROUND] Lỗi tự động đồng bộ nền: {e}")
#                 time.sleep(SYNC_ERROR_BACKOFF_SECONDS)
                

#     thread = threading.Thread(target=run_loop, daemon=True)
#     thread.start()
from apscheduler.schedulers.background import BackgroundScheduler

# Bạn có thể giữ lại các hằng số này
SYNC_INTERVAL_SECONDS = 120

def start_background_sync():
    def run_sync_job():
        with app.app_context():
            try:
                success, msg = run_sync_process(username="Auto Scheduler")
                print(f"--- [BACKGROUND SYNC] {msg} ---")
            except Exception as e:
                print(f"[BACKGROUND] Lỗi tự động đồng bộ nền: {e}")

    scheduler = BackgroundScheduler(daemon=True)
    
    # Thiết lập chạy vòng lặp mỗi 120 giây
    scheduler.add_job(run_sync_job, 'interval', seconds=SYNC_INTERVAL_SECONDS)
    scheduler.start()

@app.route("/admin/sync-sheet", methods=["POST"])
@admin_required
def sync_sheet():
    success, message = run_sync_process(username=session.get('username'))
    if success:
        flash(f"Đồng bộ Google Sheets thành công! {message}", "success")
    else:
        flash(f"Lỗi khi đồng bộ từ Google Sheets: {message}", "danger")
    return redirect(url_for('admin_panel'))

@app.route("/api/get-home-data")
def get_home_data():
    vung = session.get('vung', 'Cà Mau')
    
    st_cm = Setting.query.filter_by(key='last_updated_cm').first()
    su_cm = Setting.query.filter_by(key='last_user_cm').first()
    
    st_bl = Setting.query.filter_by(key='last_updated_bl').first()
    su_bl = Setting.query.filter_by(key='last_user_bl').first()

    if vung == 'Bạc Liêu':
        current_time = st_bl.value if st_bl else "Chưa cập nhật"
        current_user = su_bl.value if su_bl else "Hệ thống"
    else:
        current_time = st_cm.value if st_cm else "Chưa cập nhật"
        current_user = su_cm.value if su_cm else "Hệ thống"

    danh_sach_xe = Xe.query.all()
    data = []
    for xe in danh_sach_xe:
        gia_info = lay_gia_theo_vung(xe, vung)
        khuyen_mai = lay_khuyen_mai_ap_dung_cho_xe(xe.ten_xe)
        data.append({
            "id": xe.id,
            "ten_xe": xe.ten_xe,
            "loai_xe": xe.loai_xe,
            "khuyen_mai": khuyen_mai,
            "co_khuyen_mai": bool(khuyen_mai['cty'] or khuyen_mai['honda']),
            "gia_hien_thi": gia_info['gia_hien_thi'],
            "gia_cao": gia_info['gia_cao'],
            "gia_trung": gia_info['gia_trung'],
            "gia_thap": gia_info['gia_thap'],
            "gia_giay_to_phuong": gia_info['gia_gt_phuong'],
            "gia_giay_to_xa": gia_info['gia_gt_xa'],
            # Giá làm giấy tờ theo TỪNG vùng (để FE cho phép chọn vùng khác vùng đăng nhập)
            "gia_giay_to_phuong_cm": xe.gia_gt_phuong_cm or 0,
            "gia_giay_to_xa_cm": xe.gia_gt_xa_cm or 0,
            "gia_giay_to_phuong_bl": xe.gia_gt_phuong_bl or 0,
            "gia_giay_to_xa_bl": xe.gia_gt_xa_bl or 0,
            # MỚI: giá giấy tờ theo từng khu vực nhỏ Bạc Liêu (người dùng tự chọn ở FE)
            "gia_giay_to_khu_vuc_nho_bl": lay_gia_giay_to_khu_vuc_nho_bl(xe.id) if vung == 'Bạc Liêu' else [],
            "mau_xe": [mau.to_dict(vung) for mau in xe.mau_xe]
        })
        
    return jsonify({
        "success": True, 
        "vung": vung,
        "last_updated": current_time,
        "last_updated_by": current_user,
        "vungs": [
            {
                "vung": "Cà Mau",
                "last_updated": st_cm.value if st_cm else "Chưa cập nhật",
                "last_updated_by": su_cm.value if su_cm else "Hệ thống"
            },
            {
                "vung": "Bạc Liêu",
                "last_updated": st_bl.value if st_bl else "Chưa cập nhật",
                "last_updated_by": su_bl.value if su_bl else "Hệ thống"
            }
        ],
        "data": data
    })

@app.route("/admin/api/data", methods=["GET"])
@admin_required
def get_admin_data():
    try:
        danh_sach_xe = Xe.query.all()
        data = []
        for xe in danh_sach_xe:
            mau_list = []
            for mau in xe.mau_xe:
                mau_list.append({
                    "id": mau.id,
                    "ten_mau": mau.ten_mau,
                    "ns1": mau.ns1 or 0, "ns2": mau.ns2 or 0, "ns3": mau.ns3 or 0,
                    "ns4": mau.ns4 or 0, "ns5": mau.ns5 or 0, "nsm1": mau.nsm1 or 0,
                    "chenh_lech_cm": mau.chenh_lech_cm or 0,
                    "chenh_lech_bl": mau.chenh_lech_bl or 0,
                })
            data.append({
                "id": xe.id, "ten_xe": xe.ten_xe, "loai_xe": xe.loai_xe, "mau_xe": mau_list
            })
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/admin/users")
@admin_required
def manage_users():
    users = User.query.order_by(User.id.asc()).all()
    return render_template('manage_users.html', users=users, username=session.get('username'))

# --- ROUTE ĐĂNG KÝ CÔNG KHAI (TỪ TRANG LOGIN - TRẠNG THÁI CHỜ DUYỆT) ---
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")
        ho_ten = request.form.get("ho_ten", "").strip()
        bo_phan = request.form.get("bo_phan", "").strip()
        khu_vuc = request.form.get("khu_vuc", "Cà Mau")
        
        if confirm_password and password != confirm_password:
            flash("Mật khẩu xác nhận không trùng khớp!", "danger")
            return redirect(url_for('register'))

        if User.query.filter_by(username=username).first():
            flash(f"Tài khoản '{username}' đã tồn tại!", "danger")
            return redirect(url_for('register'))

        # Đăng ký mới mặc định là 'user' và trạng thái 'pending' (chờ duyệt)
        new_user = User(
            username=username, 
            password=generate_password_hash(password), 
            ho_ten=ho_ten,
            role="user", 
            bo_phan=bo_phan, 
            khu_vuc=khu_vuc,
            trang_thai="pending"
        )
        db.session.add(new_user)
        db.session.commit()
        
        flash("Đăng ký thành công! Vui lòng chờ quản trị viên duyệt tài khoản trước khi đăng nhập.", "success")
        return redirect(url_for('login'))
        
    return render_template("register.html")

# --- ROUTE ADMIN THÊM TÀI KHOẢN TRỰC TIẾP ---
@app.route("/admin/register", methods=["GET", "POST"])
@admin_required
def admin_add_user():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")
        ho_ten = request.form.get("ho_ten", "").strip()
        bo_phan = request.form.get("bo_phan", "").strip()
        khu_vuc = request.form.get("khu_vuc", "Cà Mau")
        
        is_admin = request.form.get("is_admin")
        role = "admin" if is_admin == "yes" else "user"

        if confirm_password and password != confirm_password:
            flash("Mật khẩu xác nhận không trùng khớp!", "danger")
            return redirect(url_for('admin_add_user'))

        if User.query.filter_by(username=username).first():
            flash(f"Tài khoản '{username}' đã tồn tại!", "danger")
            return redirect(url_for('admin_add_user'))

        new_user = User(
            username=username, 
            password=generate_password_hash(password), 
            ho_ten=ho_ten,
            role=role, 
            bo_phan=bo_phan, 
            khu_vuc=khu_vuc,
            trang_thai="approved" # Tài khoản do admin tạo được duyệt sẵn
        )
        db.session.add(new_user)
        db.session.commit()
        
        flash("Tạo tài khoản người dùng thành công!", "success")
        return redirect(url_for('manage_users'))
        
    return render_template("register.html", username=session.get('username'))

# --- DUYỆT HOẶC TỪ CHỐI TÀI KHOẢN (DÀNH CHO ADMIN) ---
@app.route("/admin/users/approve/<int:id>")
@admin_required
def approve_user(id):  # <--- Đổi 'user_id' thành 'id' cho khớp với <int:id> ở trên
    user = User.query.get_or_404(id)
    user.trang_thai = 'approved'
    db.session.commit()
    
    return redirect(url_for('admin_panel'))
@app.route("/admin/users/reject/<int:id>")
@admin_required
def reject_user(id):
    user = db.get_or_404(User, id)
    user.trang_thai = "rejected"
    db.session.commit()
    flash(f"Đã từ chối tài khoản {user.username}!", "warning")
    return redirect(url_for('manage_users'))

@app.route("/admin/users/edit/<int:id>", methods=["POST"])
@admin_required
def edit_user(id):
    user = db.get_or_404(User, id)
    new_password = request.form.get("password")
    if new_password:
        user.password = generate_password_hash(new_password)
        
    user.ho_ten = request.form.get("ho_ten", "").strip()
    user.bo_phan = request.form.get("bo_phan")
    user.khu_vuc = request.form.get("khu_vuc")
    user.role = 'admin' if request.form.get("is_admin") == 'yes' else 'user'
    
    db.session.commit()
    flash(f"Cập nhật tài khoản {user.username} thành công!", "success")
    return redirect(url_for('manage_users'))

@app.route("/admin/users/delete/<int:id>")
@admin_required
def delete_user(id):
    user = db.get_or_404(User, id)
    if user.username == session.get('username'):
        flash("Không thể xóa tài khoản của chính bạn!", "danger")
        return redirect(url_for('manage_users'))
        
    db.session.delete(user)
    db.session.commit()
    flash("Đã xóa tài khoản thành công!", "success")
    return redirect(url_for('manage_users'))

@app.route("/admin/add", methods=["POST"])
@admin_required
def add_xe():
    ten_xe_nhap = request.form.get("ten_xe", "").strip()
    if Xe.query.filter_by(ten_xe=ten_xe_nhap).first():
        flash(f"Lỗi: Tên xe '{ten_xe_nhap}' đã tồn tại!", "danger")
        return redirect(url_for('admin_panel'))

    loai_xe_nhap = request.form.get("loai_xe", "").strip()
    if not loai_xe_nhap or loai_xe_nhap == "Chưa phân loại":
        loai_xe_nhap = tu_dong_phan_loai(ten_xe_nhap)

    try:
        new_xe = Xe(
            loai_xe=loai_xe_nhap, 
            ten_xe=ten_xe_nhap,
            phien_ban=request.form.get("phien_ban"), 
            gia_cm_thap=safe_float(request.form.get("gia_cm_thap")),
            gia_cm_trung=safe_float(request.form.get("gia_cm_trung")),
            gia_cm_cao=safe_float(request.form.get("gia_cm_cao")),
            gia_bl_thap=safe_float(request.form.get("gia_bl_thap")),
            gia_bl_trung=safe_float(request.form.get("gia_bl_trung")),
            gia_bl_cao=safe_float(request.form.get("gia_bl_cao")),
            gia_gt_phuong_cm=safe_float(request.form.get("gia_gt_phuong_cm")),
            gia_gt_xa_cm=safe_float(request.form.get("gia_gt_xa_cm")),
            gia_gt_phuong_bl=safe_float(request.form.get("gia_gt_phuong_bl")),
            gia_gt_xa_bl=safe_float(request.form.get("gia_gt_xa_bl")),
            ns1=int(request.form.get("ns1") or 0), ns2=int(request.form.get("ns2") or 0),
            ns3=int(request.form.get("ns3") or 0), ns4=int(request.form.get("ns4") or 0),
            ns5=int(request.form.get("ns5") or 0), nsm1=int(request.form.get("nsm1") or 0),
            hinh_anh=save_image(request.files.get('hinh_anh'))
        )
        db.session.add(new_xe)
        db.session.flush()
        
        ten_maus = request.form.getlist("ten_mau[]")
        chenh_lechs_cm = request.form.getlist("chenh_lech_cm[]")
        chenh_lechs_bl = request.form.getlist("chenh_lech_bl[]")
        anh_maus = request.files.getlist("hinh_anh_mau[]")
        
        for i in range(len(ten_maus)):
            if ten_maus[i].strip():
                c_cm = safe_float(chenh_lechs_cm[i]) if (i < len(chenh_lechs_cm) and chenh_lechs_cm[i]) else 0
                c_bl = safe_float(chenh_lechs_bl[i]) if (i < len(chenh_lechs_bl) and chenh_lechs_bl[i]) else 0
                db.session.add(XeMau(
                    xe_id=new_xe.id, ten_mau=ten_maus[i].strip(), 
                    chenh_lech_cm=c_cm, chenh_lech_bl=c_bl,
                    hinh_anh_mau=save_image(anh_maus[i] if i < len(anh_maus) else None)
                ))
                
        db.session.commit()
        
        if new_xe.gia_cm_cao > 0:
            cap_nhat_thoi_gian_dong_bo("Cà Mau", session.get('username'))
        if new_xe.gia_bl_cao > 0:
            cap_nhat_thoi_gian_dong_bo("Bạc Liêu", session.get('username'))

        flash("Thêm xe mới thành công!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Có lỗi xảy ra: {str(e)}", "danger")

    return redirect(url_for('admin_panel'))

@app.route("/admin/edit/<int:id>", methods=["POST"])
@admin_required
def edit_xe(id):
    xe = db.get_or_404(Xe, id)
    ten_xe_moi = request.form.get("ten_xe", "").strip()
    if ten_xe_moi and ten_xe_moi != xe.ten_xe and Xe.query.filter_by(ten_xe=ten_xe_moi).first():
        flash(f"Lỗi: Tên xe '{ten_xe_moi}' đã tồn tại!", "danger")
        return redirect(url_for('admin_panel'))

    loai_xe_nhap = request.form.get("loai_xe", "").strip()
    if not loai_xe_nhap or loai_xe_nhap == "Chưa phân loại":
        loai_xe_nhap = tu_dong_phan_loai(ten_xe_moi if ten_xe_moi else xe.ten_xe)

    old_cm_cao = xe.gia_cm_cao
    old_bl_cao = xe.gia_bl_cao
    old_cm_trung = xe.gia_cm_trung
    old_bl_trung = xe.gia_bl_trung
    old_cm_thap = xe.gia_cm_thap
    old_bl_thap = xe.gia_bl_thap

    try:
        xe.loai_xe = loai_xe_nhap
        xe.ten_xe = ten_xe_moi
        xe.phien_ban = request.form.get("phien_ban")
        xe.gia_cm_thap = safe_float(request.form.get("gia_cm_thap"))
        xe.gia_cm_trung = safe_float(request.form.get("gia_cm_trung"))
        xe.gia_cm_cao = safe_float(request.form.get("gia_cm_cao"))
        xe.gia_bl_thap = safe_float(request.form.get("gia_bl_thap"))
        xe.gia_bl_trung = safe_float(request.form.get("gia_bl_trung"))
        xe.gia_bl_cao = safe_float(request.form.get("gia_bl_cao"))
        xe.gia_gt_phuong_cm = safe_float(request.form.get("gia_gt_phuong_cm"))
        xe.gia_gt_xa_cm = safe_float(request.form.get("gia_gt_xa_cm"))
        xe.gia_gt_phuong_bl = safe_float(request.form.get("gia_gt_phuong_bl"))
        xe.gia_gt_xa_bl = safe_float(request.form.get("gia_gt_xa_bl"))
        xe.ns1 = int(request.form.get("ns1") or 0); xe.ns2 = int(request.form.get("ns2") or 0)
        xe.ns3 = int(request.form.get("ns3") or 0); xe.ns4 = int(request.form.get("ns4") or 0)
        xe.ns5 = int(request.form.get("ns5") or 0); xe.nsm1 = int(request.form.get("nsm1") or 0)
        
        if request.files.get('hinh_anh') and request.files.get('hinh_anh').filename != '':
            xe.hinh_anh = save_image(request.files.get('hinh_anh'))
        
        for mau in xe.mau_xe:
            mau.ten_mau = request.form.get(f"edit_ten_mau_{mau.id}", mau.ten_mau)
            mau.chenh_lech_cm = safe_float(request.form.get(f"edit_chenh_lech_cm_{mau.id}"))
            mau.chenh_lech_bl = safe_float(request.form.get(f"edit_chenh_lech_bl_{mau.id}"))
            if request.files.get(f"edit_hinh_anh_mau_{mau.id}") and request.files.get(f"edit_hinh_anh_mau_{mau.id}").filename != '': 
                mau.hinh_anh_mau = save_image(request.files.get(f"edit_hinh_anh_mau_{mau.id}"))
            
        new_tens = request.form.getlist("new_ten_mau[]")
        new_cms = request.form.getlist("new_chenh_lech_cm[]")
        new_bls = request.form.getlist("new_chenh_lech_bl[]")
        new_anhs = request.files.getlist("new_hinh_anh_mau[]")
        
        for i in range(len(new_tens)):
            if new_tens[i].strip():
                c_cm = safe_float(new_cms[i]) if (i < len(new_cms) and new_cms[i]) else 0
                c_bl = safe_float(new_bls[i]) if (i < len(new_bls) and new_bls[i]) else 0
                db.session.add(XeMau(
                    xe_id=xe.id, ten_mau=new_tens[i].strip(), 
                    chenh_lech_cm=c_cm, chenh_lech_bl=c_bl,
                    hinh_anh_mau=save_image(new_anhs[i] if i < len(new_anhs) else None)
                ))

        db.session.commit()

        has_price_change = (
            xe.gia_cm_cao != old_cm_cao or xe.gia_cm_trung != old_cm_trung or xe.gia_cm_thap != old_cm_thap or
            xe.gia_bl_cao != old_bl_cao or xe.gia_bl_trung != old_bl_trung or xe.gia_bl_thap != old_bl_thap
        )

        if has_price_change:
            try:
                sql_log = text("""
                    INSERT INTO history_logs (username, action, target_id, old_value, new_value) 
                    VALUES (:username, :action, :target_id, :old_val, :new_val)
                """)
                db.session.execute(sql_log, {
                    'username': session.get('username', 'Admin'),
                    'action': 'Cập nhật giá qua Form Sửa',
                    'target_id': f"{xe.ten_xe} (ID: {xe.id})",
                    'old_val': f"CM [Cao: {old_cm_cao}, Trung: {old_cm_trung}, Thấp: {old_cm_thap}] | BL [Cao: {old_bl_cao}, Trung: {old_bl_trung}, Thấp: {old_bl_thap}]",
                    'new_val': f"CM [Cao: {xe.gia_cm_cao}, Trung: {xe.gia_cm_trung}, Thấp: {xe.gia_cm_thap}] | BL [Cao: {xe.gia_bl_cao}, Trung: {xe.gia_bl_trung}, Thấp: {xe.gia_bl_thap}]"
                })
                db.session.commit()
            except Exception as e:
                print("Lỗi ghi log lịch sử:", e)

        if xe.gia_cm_cao != old_cm_cao or xe.gia_cm_trung != old_cm_trung or xe.gia_cm_thap != old_cm_thap:
            cap_nhat_thoi_gian_dong_bo("Cà Mau", session.get('username'))
        if xe.gia_bl_cao != old_bl_cao or xe.gia_bl_trung != old_bl_trung or xe.gia_bl_thap != old_bl_thap:
            cap_nhat_thoi_gian_dong_bo("Bạc Liêu", session.get('username'))

    except Exception as e:
        db.session.rollback()
        flash(f"Có lỗi xảy ra: {str(e)}", "danger")

    return redirect(url_for('admin_panel'))

@app.route("/admin/delete/<int:id>", methods=["GET"])
@admin_required
def delete_xe(id):
    try:
        db.session.delete(db.get_or_404(Xe, id))
        db.session.commit()
        flash("Đã xóa xe thành công!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Không thể xóa xe: {str(e)}", "danger")
    return redirect(url_for('admin_panel'))

@app.route("/admin/delete-mau/<int:id>", methods=["GET"])
@admin_required
def delete_mau(id):
    try:
        db.session.delete(db.get_or_404(XeMau, id))
        db.session.commit()
        flash("Đã xóa màu thành công!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Không thể xóa màu: {str(e)}", "danger")
    return redirect(url_for('admin_panel'))

@app.route("/admin/import", methods=["POST"])
@admin_required
def import_excel():
    file = request.files.get('file_excel')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash("Vui lòng chọn tập tin Excel hợp lệ (.xlsx, .xls)!", "danger")
        return redirect(url_for('admin_panel'))

    try:
        df = pd.read_excel(file).fillna(0)
        df.columns = df.columns.str.strip().str.lower()
        
        so_luong_them = 0
        so_luong_cap_nhat = 0
        has_price_changed_cm = False
        has_price_changed_bl = False

        for _, row in df.iterrows():
            ten_xe_excel = str(row.get('ten_xe', '')).strip()
            if not ten_xe_excel or ten_xe_excel == '0': 
                continue
            
            xe = Xe.query.filter_by(ten_xe=ten_xe_excel).first()
            loai_xe_excel = str(row.get('loai_xe', '')).strip()
            if not loai_xe_excel or loai_xe_excel == '0' or loai_xe_excel == 'Chưa phân loại':
                loai_xe_excel = tu_dong_phan_loai(ten_xe_excel)
            
            if xe:
                if loai_xe_excel: xe.loai_xe = loai_xe_excel
                if row.get('phien_ban'): xe.phien_ban = str(row.get('phien_ban')).strip()
                
                g_cm_cao = safe_float(row.get('gia_cm_cao'), xe.gia_cm_cao)
                g_cm_trung = safe_float(row.get('gia_cm_trung'), xe.gia_cm_trung)
                g_cm_thap = safe_float(row.get('gia_cm_thap'), xe.gia_cm_thap)
                
                g_bl_cao = safe_float(row.get('gia_bl_cao'), xe.gia_bl_cao)
                g_bl_trung = safe_float(row.get('gia_bl_trung'), xe.gia_bl_trung)
                g_bl_thap = safe_float(row.get('gia_bl_thap'), xe.gia_bl_thap)

                if xe.gia_cm_cao != g_cm_cao or xe.gia_cm_trung != g_cm_trung or xe.gia_cm_thap != g_cm_thap:
                    has_price_changed_cm = True
                if xe.gia_bl_cao != g_bl_cao or xe.gia_bl_trung != g_bl_trung or xe.gia_bl_thap != g_bl_thap:
                    has_price_changed_bl = True

                xe.gia_cm_thap = g_cm_thap
                xe.gia_cm_trung = g_cm_trung
                xe.gia_cm_cao = g_cm_cao
                xe.gia_bl_thap = g_bl_thap
                xe.gia_bl_trung = g_bl_trung
                xe.gia_bl_cao = g_bl_cao

                xe.gia_gt_phuong_cm = safe_float(row.get('gia_gt_phuong_cm'), xe.gia_gt_phuong_cm)
                xe.gia_gt_xa_cm = safe_float(row.get('gia_gt_xa_cm'), xe.gia_gt_xa_cm)
                xe.gia_gt_phuong_bl = safe_float(row.get('gia_gt_phuong_bl'), xe.gia_gt_phuong_bl)
                xe.gia_gt_xa_bl = safe_float(row.get('gia_gt_xa_bl'), xe.gia_gt_xa_bl)
                xe.ns1 = safe_int(row.get('ns1'), xe.ns1); xe.ns2 = safe_int(row.get('ns2'), xe.ns2)
                xe.ns3 = safe_int(row.get('ns3'), xe.ns3); xe.ns4 = safe_int(row.get('ns4'), xe.ns4)
                xe.ns5 = safe_int(row.get('ns5'), xe.ns5); xe.nsm1 = safe_int(row.get('nsm1'), xe.nsm1)
                so_luong_cap_nhat += 1
            else:
                xe = Xe(
                    loai_xe=loai_xe_excel, ten_xe=ten_xe_excel,
                    phien_ban=str(row.get('phien_ban', '')).strip(),
                    gia_cm_thap=safe_float(row.get('gia_cm_thap')),
                    gia_cm_trung=safe_float(row.get('gia_cm_trung')),
                    gia_cm_cao=safe_float(row.get('gia_cm_cao')),
                    gia_bl_thap=safe_float(row.get('gia_bl_thap')),
                    gia_bl_trung=safe_float(row.get('gia_bl_trung')),
                    gia_bl_cao=safe_float(row.get('gia_bl_cao')),
                    gia_gt_phuong_cm=safe_float(row.get('gia_gt_phuong_cm')),
                    gia_gt_xa_cm=safe_float(row.get('gia_gt_xa_cm')),
                    gia_gt_phuong_bl=safe_float(row.get('gia_gt_phuong_bl')),
                    gia_gt_xa_bl=safe_float(row.get('gia_gt_xa_bl')),
                    ns1=safe_int(row.get('ns1')), ns2=safe_int(row.get('ns2')),
                    ns3=safe_int(row.get('ns3')), ns4=safe_int(row.get('ns4')),
                    ns5=safe_int(row.get('ns5')), nsm1=safe_int(row.get('nsm1'))
                )
                db.session.add(xe)
                db.session.flush()
                so_luong_them += 1
                if xe.gia_cm_cao > 0 or xe.gia_cm_trung > 0 or xe.gia_cm_thap > 0: has_price_changed_cm = True
                if xe.gia_bl_cao > 0 or xe.gia_bl_trung > 0 or xe.gia_bl_thap > 0: has_price_changed_bl = True

            ten_mau_excel = str(row.get('ten_mau', '')).strip()
            if ten_mau_excel and ten_mau_excel != '0':
                mau_existing = XeMau.query.filter_by(xe_id=xe.id, ten_mau=ten_mau_excel).first()
                cl_cm = safe_float(row.get('chenh_lech_cm'))
                cl_bl = safe_float(row.get('chenh_lech_bl'))
                m_ns1 = safe_int(row.get('mau_ns1', row.get('ns1', 0)))
                m_ns2 = safe_int(row.get('mau_ns2', row.get('ns2', 0)))
                m_ns3 = safe_int(row.get('mau_ns3', row.get('ns3', 0)))
                m_ns4 = safe_int(row.get('mau_ns4', row.get('ns4', 0)))
                m_ns5 = safe_int(row.get('mau_ns5', row.get('ns5', 0)))
                m_nsm1 = safe_int(row.get('mau_nsm1', row.get('nsm1', 0)))

                if mau_existing:
                    mau_existing.chenh_lech_cm = cl_cm; mau_existing.chenh_lech_bl = cl_bl
                    mau_existing.ns1 = m_ns1; mau_existing.ns2 = m_ns2; mau_existing.ns3 = m_ns3
                    mau_existing.ns4 = m_ns4; mau_existing.ns5 = m_ns5; mau_existing.nsm1 = m_nsm1
                else:
                    db.session.add(XeMau(
                        xe_id=xe.id, ten_mau=ten_mau_excel,
                        chenh_lech_cm=cl_cm, chenh_lech_bl=cl_bl,
                        ns1=m_ns1, ns2=m_ns2, ns3=m_ns3, ns4=m_ns4, ns5=m_ns5, nsm1=m_nsm1
                    ))

        db.session.commit()
        
        if has_price_changed_cm:
            cap_nhat_thoi_gian_dong_bo("Cà Mau", session.get('username'))
        if has_price_changed_bl:
            cap_nhat_thoi_gian_dong_bo("Bạc Liêu", session.get('username'))

        flash(f"Thao tác thành công! Thêm mới: {so_luong_them} xe, Cập nhật: {so_luong_cap_nhat} xe.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi khi xử lý file Excel: {str(e)}", "danger")

    return redirect(url_for('admin_panel'))

@app.route('/api/sync-inventory', methods=['POST'])
def sync_inventory_api():
    try:
        data = request.json
        store_code = data.get('store_code')
        rows = data.get('rows', [])
        
        if not store_code or not rows:
            return jsonify({"status": "error", "message": "Thiếu dữ liệu"}), 400
            
        all_xe = {xe.ten_xe: xe for xe in Xe.query.all()}
        all_mau = {(m.xe_id, m.ten_mau): m for m in XeMau.query.all()}
        
        for item in rows:
            ten_xe = item.get('ten_xe')
            ten_mau = item.get('ten_mau')
            ton_cuoi = item.get('ton_cuoi', 0)
            
            if not ten_xe: continue
            
            if ten_xe not in all_xe:
                loai_xe_auto = tu_dong_phan_loai(ten_xe)
                xe_obj = Xe(ten_xe=ten_xe, loai_xe=loai_xe_auto)
                db.session.add(xe_obj)
                db.session.flush() 
                all_xe[ten_xe] = xe_obj
            else:
                xe_obj = all_xe[ten_xe]
                if not xe_obj.loai_xe or xe_obj.loai_xe == "Chưa phân loại":
                    xe_obj.loai_xe = tu_dong_phan_loai(ten_xe)

            key_mau = (xe_obj.id, ten_mau)
            if key_mau not in all_mau:
                mau_obj = XeMau(xe_id=xe_obj.id, ten_mau=ten_mau)
                db.session.add(mau_obj)
                db.session.flush()
                all_mau[key_mau] = mau_obj
            else:
                mau_obj = all_mau[key_mau]

            if store_code == "NS1": mau_obj.ns1 = ton_cuoi
            elif store_code == "NS2": mau_obj.ns2 = ton_cuoi
            elif store_code == "NS3": mau_obj.ns3 = ton_cuoi
            elif store_code == "NS4": mau_obj.ns4 = ton_cuoi
            elif store_code == "NS5": mau_obj.ns5 = ton_cuoi
            elif store_code == "NSM1": mau_obj.nsm1 = ton_cuoi
                
        db.session.commit()
        return jsonify({"status": "success", "message": f"Đồng bộ thành công kho {store_code}"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500

def is_bac_lieu(khu_vuc):
    kv = (khu_vuc or '').lower()
    return 'bạc liêu' in kv or 'bac lieu' in kv

def lay_gia_theo_vung(xe, khu_vuc_user):
    if is_bac_lieu(khu_vuc_user):
        return {
            'gia_hien_thi': getattr(xe, 'gia_bl_cao', 0) or 0,
            'gia_gt_phuong': getattr(xe, 'gia_gt_phuong_bl', 0) or 0,
            'gia_gt_xa': getattr(xe, 'gia_gt_xa_bl', 0) or 0,
            'gia_thap': getattr(xe, 'gia_bl_thap', 0) or 0,
            'gia_trung': getattr(xe, 'gia_bl_trung', 0) or 0,
            'gia_cao': getattr(xe, 'gia_bl_cao', 0) or 0
        }
    else:
        return {
            'gia_hien_thi': getattr(xe, 'gia_cm_cao', 0) or 0,
            'gia_gt_phuong': getattr(xe, 'gia_gt_phuong_cm', 0) or 0,
            'gia_gt_xa': getattr(xe, 'gia_gt_xa_cm', 0) or 0,
            'gia_thap': getattr(xe, 'gia_cm_thap', 0) or 0,
            'gia_trung': getattr(xe, 'gia_cm_trung', 0) or 0,
            'gia_cao': getattr(xe, 'gia_cm_cao', 0) or 0
        }

def lay_gia_giay_to_khu_vuc_nho_bl(xe_id):
    """
    Trả về danh sách giá giấy tờ của 1 xe theo TỪNG khu vực nhỏ Bạc Liêu,
    gom nhóm theo khu vực lớn (Nam Sương 4 / Nam Sương 2 / Nam Sương 5 - NSM1).
    Dùng để người dùng tự chọn vùng giấy tờ ở tinh_gia_nhap.html.
    """
    ds_gia = {g.khu_vuc_nho_id: (g.gia or 0) for g in GiaGiayToXeBL.query.filter_by(xe_id=xe_id).all()}
    ds_khu_vuc_lon = KhuVucLonBL.query.order_by(KhuVucLonBL.thu_tu).all()

    ket_qua = []
    for kvl in ds_khu_vuc_lon:
        ket_qua.append({
            'khu_vuc_lon_id': kvl.id,
            'ma_khu_vuc': kvl.ma_khu_vuc,
            'ten_khu_vuc_lon': kvl.ten_khu_vuc,
            'khu_vuc_nho': [
                {
                    'id': kvn.id,
                    'ten_khu_vuc_nho': kvn.ten_khu_vuc_nho,
                    'gia': ds_gia.get(kvn.id, 0)
                } for kvn in sorted(kvl.khu_vuc_nho, key=lambda x: x.thu_tu)
            ]
        })
    return ket_qua

def format_xe_data_home(xe, khu_vuc_user):
    is_bl = 'bạc liêu' in (khu_vuc_user or '').lower()
    gia_thap = xe.gia_bl_thap if is_bl else xe.gia_cm_thap
    gia_trung = xe.gia_bl_trung if is_bl else xe.gia_cm_trung
    gia_cao = xe.gia_bl_cao if is_bl else xe.gia_cm_cao
    # Cà Mau: giữ nguyên logic phường/xã cố định như cũ.
    # Bạc Liêu: giá giấy tờ không còn cố định 1 mức, người dùng sẽ tự chọn khu vực nhỏ
    # trong 3 khu vực lớn (xem 'gia_giay_to_khu_vuc_nho_bl' bên dưới), nên các trường
    # gia_giay_to_phuong/xa chỉ còn ý nghĩa với Cà Mau.
    gia_giay_to_phuong = xe.gia_gt_phuong_cm if not is_bl else None
    gia_giay_to_xa = xe.gia_gt_xa_cm if not is_bl else None

    khuyen_mai = lay_khuyen_mai_ap_dung_cho_xe(xe.ten_xe)

    return {
        'id': xe.id,
        'loai_xe': xe.loai_xe,
        'ten_xe': xe.ten_xe,
        'phien_ban': xe.phien_ban,
        'gia_thap': gia_thap,
        'gia_trung': gia_trung,
        'gia_cao': gia_cao,
        'gia_giay_to_phuong': gia_giay_to_phuong,
        'gia_giay_to_xa': gia_giay_to_xa,
        'gia_giay_to_phuong_ca_mau': xe.gia_gt_phuong_cm,
        'gia_giay_to_xa_ca_mau': xe.gia_gt_xa_cm,
        # Giữ 2 trường cũ này để không phá vỡ chỗ nào còn tham chiếu (không còn dùng cho tính giá BL nữa)
        'gia_giay_to_phuong_bac_lieu': xe.gia_gt_phuong_bl,
        'gia_giay_to_xa_bac_lieu': xe.gia_gt_xa_bl,
        # MỚI: giá giấy tờ theo từng khu vực nhỏ Bạc Liêu, chỉ cần điền khi is_bl=True
        'gia_giay_to_khu_vuc_nho_bl': lay_gia_giay_to_khu_vuc_nho_bl(xe.id) if is_bl else [],
        'hinh_anh': xe.hinh_anh,
        'mau_xe': [mau.to_dict(khu_vuc_user) for mau in xe.mau_xe],
        'ns1': xe.ns1, 'ns2': xe.ns2, 'ns3': xe.ns3,
        'ns4': xe.ns4, 'ns5': xe.ns5, 'nsm1': xe.nsm1,
        # MỚI: khuyến mãi đang áp dụng cho xe này (chỉ gồm CTKM còn hạn, tự động ẩn khi hết hạn)
        'khuyen_mai': khuyen_mai,
        'co_khuyen_mai': bool(khuyen_mai['cty'] or khuyen_mai['honda'])
    }

@app.route('/cong-cu-tinh-gia')
def cong_cu_tinh_gia():
    return render_template('tinh_gia_nhap.html')

@app.route('/admin/gia-giay-to-bl')
@admin_required
def admin_gia_giay_to_bl_page():
    """Trang quản trị: nhập giá giấy tờ theo từng khu vực nhỏ Bạc Liêu cho từng xe."""
    danh_sach_xe = Xe.query.order_by(get_order_priority(), Xe.ten_xe.asc()).all()
    khu_vuc_lon = KhuVucLonBL.query.order_by(KhuVucLonBL.thu_tu).all()
    khu_vuc_lon_data = [{
        'id': kvl.id, 'ma_khu_vuc': kvl.ma_khu_vuc, 'ten_khu_vuc': kvl.ten_khu_vuc,
        'khu_vuc_nho': [{'id': n.id, 'ten_khu_vuc_nho': n.ten_khu_vuc_nho} for n in sorted(kvl.khu_vuc_nho, key=lambda x: x.thu_tu)]
    } for kvl in khu_vuc_lon]
    xe_data = [{
        'id': xe.id, 'ten_xe': xe.ten_xe, 'phien_ban': xe.phien_ban,
        'gia_theo_khu_vuc_nho': lay_gia_giay_to_khu_vuc_nho_bl(xe.id)
    } for xe in danh_sach_xe]
    return render_template('admin_gia_giay_to_bl.html', khu_vuc_lon=khu_vuc_lon_data, danh_sach_xe=xe_data)

def _chuan_hoa_ten(s):
    """Chuẩn hóa tên phường/xã để so khớp: bỏ khoảng trắng thừa, thường hóa, giữ dấu tiếng Việt."""
    return unicodedata.normalize('NFC', str(s or '')).strip().lower()

@app.route('/admin/api/import-gia-giay-to-bl', methods=['POST'])
@admin_required
def admin_import_gia_giay_to_bl():
    """
    Nhập giá giấy tờ Bạc Liêu từ file Excel dạng:
    - Cột A: 'ten_xe' (phải khớp CHÍNH XÁC với Xe.ten_xe trong CSDL)
    - Các cột B trở đi: mỗi cột là 1 phường/xã, tên phường/xã ở dòng tiêu đề (dòng 1).
      Các cột thuộc cùng 1 khu vực nhỏ sẽ được nhận diện qua tên phường/xã khớp với
      dữ liệu đã seed sẵn (không bắt buộc phải tô cùng màu, màu chỉ để người nhập liệu dễ nhìn).
    Body form-data: file (bắt buộc), ma_khu_vuc (vd 'NS4', mặc định 'NS4')
    """
    file = request.files.get('file')
    if not file or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Vui lòng chọn tập tin Excel hợp lệ (.xlsx, .xls)!'}), 400

    ma_khu_vuc = (request.form.get('ma_khu_vuc') or 'NS4').strip()
    kvl = KhuVucLonBL.query.filter_by(ma_khu_vuc=ma_khu_vuc).first()
    if not kvl:
        return jsonify({'success': False, 'message': f'Không tìm thấy khu vực lớn có mã "{ma_khu_vuc}".'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # Đọc dòng tiêu đề (dòng 1): cột B trở đi = tên phường/xã
        header_cols = []  # [(col_index, ten_phuong_xa)]
        for col_idx in range(2, ws.max_column + 1):
            ten_header = ws.cell(row=1, column=col_idx).value
            if ten_header and str(ten_header).strip():
                header_cols.append((col_idx, str(ten_header).strip()))

        # Khớp từng cột với 1 khu vực nhỏ đã seed (dựa trên tên phường/xã có chứa trong ten_khu_vuc_nho)
        ds_khu_vuc_nho = sorted(kvl.khu_vuc_nho, key=lambda x: x.thu_tu)
        col_to_khuvucnho = {}
        headers_khong_khop = []
        for col_idx, ten_header in header_cols:
            ten_header_norm = _chuan_hoa_ten(ten_header)
            khop = None
            for kvn in ds_khu_vuc_nho:
                if ten_header_norm in _chuan_hoa_ten(kvn.ten_khu_vuc_nho):
                    khop = kvn
                    break
            if khop:
                col_to_khuvucnho[col_idx] = khop
            else:
                headers_khong_khop.append(ten_header)

        # Đọc từng dòng dữ liệu (từ dòng 2)
        so_xe_cap_nhat = 0
        so_gia_cap_nhat = 0
        xe_khong_tim_thay = []
        for row_idx in range(2, ws.max_row + 1):
            ten_xe_excel = ws.cell(row=row_idx, column=1).value
            ten_xe_excel = str(ten_xe_excel).strip() if ten_xe_excel is not None else ''
            if not ten_xe_excel:
                continue

            xe = Xe.query.filter_by(ten_xe=ten_xe_excel).first()
            if not xe:
                xe_khong_tim_thay.append(ten_xe_excel)
                continue

            co_cap_nhat_xe_nay = False
            for col_idx, kvn in col_to_khuvucnho.items():
                gia_moi = safe_float(ws.cell(row=row_idx, column=col_idx).value, None)
                if gia_moi is None:
                    continue
                row_db = GiaGiayToXeBL.query.filter_by(xe_id=xe.id, khu_vuc_nho_id=kvn.id).first()
                if row_db:
                    if row_db.gia != gia_moi:
                        row_db.gia = gia_moi
                        so_gia_cap_nhat += 1
                        co_cap_nhat_xe_nay = True
                else:
                    db.session.add(GiaGiayToXeBL(xe_id=xe.id, khu_vuc_nho_id=kvn.id, gia=gia_moi))
                    so_gia_cap_nhat += 1
                    co_cap_nhat_xe_nay = True

            if co_cap_nhat_xe_nay:
                so_xe_cap_nhat += 1

        db.session.commit()

        if so_gia_cap_nhat:
            cap_nhat_thoi_gian_dong_bo('Bạc Liêu', session.get('username'))

        return jsonify({
            'success': True,
            'message': f'Đã cập nhật {so_gia_cap_nhat} mức giá cho {so_xe_cap_nhat} xe, khu vực {kvl.ten_khu_vuc}.',
            'khu_vuc_lon': kvl.ten_khu_vuc,
            'cot_da_khop': [{'cot': h, 'khu_vuc_nho': col_to_khuvucnho[c].ten_khu_vuc_nho} for c, h in header_cols if c in col_to_khuvucnho],
            'cot_khong_khop': headers_khong_khop,
            'xe_khong_tim_thay': xe_khong_tim_thay
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Lỗi khi xử lý file Excel: {str(e)}'}), 500

@app.route('/admin/gia-giay-to')
@admin_required
def admin_gia_giay_to_page():
    """
    Trang quản trị RIÊNG BIỆT, hợp nhất quản lý giá giấy tờ:
    - Tab Cà Mau: giá giấy tờ Phường / Xã (2 giá trị cố định cho mỗi xe).
    - Tab Bạc Liêu: giá giấy tờ theo từng khu vực nhỏ (Nam Sương 4 / Nam Sương 2 / Nam Sương 5 - NSM1).
    Toàn bộ được sửa trực tiếp trên bảng rồi lưu 1 lần bằng nút "Lưu tất cả".
    """
    danh_sach_xe_obj = Xe.query.order_by(get_order_priority(), Xe.ten_xe.asc()).all()

    khu_vuc_lon = KhuVucLonBL.query.order_by(KhuVucLonBL.thu_tu).all()
    khu_vuc_lon_data = [{
        'id': kvl.id,
        'ma_khu_vuc': kvl.ma_khu_vuc,
        'ten_khu_vuc': kvl.ten_khu_vuc,
        'khu_vuc_nho': [
            {'id': n.id, 'ten_khu_vuc_nho': n.ten_khu_vuc_nho}
            for n in sorted(kvl.khu_vuc_nho, key=lambda x: x.thu_tu)
        ]
    } for kvl in khu_vuc_lon]

    # Lấy 1 lần toàn bộ giá giấy tờ Bạc Liêu, gom theo xe để tránh truy vấn lặp lại trong vòng lặp
    gia_bl_theo_xe = {}
    for g in GiaGiayToXeBL.query.all():
        gia_bl_theo_xe.setdefault(g.xe_id, {})[g.khu_vuc_nho_id] = g.gia or 0

    danh_sach_xe = [{
        'id': xe.id,
        'ten_xe': xe.ten_xe,
        'phien_ban': xe.phien_ban,
        'loai_xe': xe.loai_xe,
        'gia_gt_phuong_cm': xe.gia_gt_phuong_cm or 0,
        'gia_gt_xa_cm': xe.gia_gt_xa_cm or 0,
        'gia_bl_map': gia_bl_theo_xe.get(xe.id, {})
    } for xe in danh_sach_xe_obj]

    return render_template('admin_gia_giay_to.html', khu_vuc_lon=khu_vuc_lon_data, danh_sach_xe=danh_sach_xe)


@app.route('/admin/api/xuat-excel-gia-giay-to')
@admin_required
def admin_xuat_excel_gia_giay_to():
    """
    Xuất toàn bộ giá giấy tờ ra 1 file Excel nhiều sheet:
    - Sheet 'Cà Mau': Tên xe, Phiên bản, Giá Giấy Tờ Phường, Giá Giấy Tờ Xã.
    - Mỗi khu vực lớn Bạc Liêu (Nam Sương 4 / Nam Sương 2 / Nam Sương 5 - NSM1) là 1 sheet riêng,
      cột là từng khu vực nhỏ thuộc khu vực lớn đó.
    """
    danh_sach_xe = Xe.query.order_by(get_order_priority(), Xe.ten_xe.asc()).all()
    khu_vuc_lon = KhuVucLonBL.query.order_by(KhuVucLonBL.thu_tu).all()

    gia_bl_theo_xe = {}
    for g in GiaGiayToXeBL.query.all():
        gia_bl_theo_xe.setdefault(g.xe_id, {})[g.khu_vuc_nho_id] = g.gia or 0

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1A56DB', end_color='1A56DB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def ke_dong_tieu_de(ws, headers):
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = 'C2' if len(headers) > 1 else 'A2'

    wb = openpyxl.Workbook()

    # --- Sheet Cà Mau ---
    ws_cm = wb.active
    ws_cm.title = 'Cà Mau'
    ke_dong_tieu_de(ws_cm, ['Tên xe', 'Phiên bản', 'Giá Giấy Tờ Phường', 'Giá Giấy Tờ Xã'])
    for xe in danh_sach_xe:
        ws_cm.append([xe.ten_xe, xe.phien_ban or '', xe.gia_gt_phuong_cm or 0, xe.gia_gt_xa_cm or 0])
    ws_cm.column_dimensions['A'].width = 34
    ws_cm.column_dimensions['B'].width = 18
    ws_cm.column_dimensions['C'].width = 20
    ws_cm.column_dimensions['D'].width = 20
    ws_cm.freeze_panes = 'A2'

    # --- Mỗi khu vực lớn Bạc Liêu là 1 sheet ---
    ten_sheet_da_dung = set()
    for kvl in khu_vuc_lon:
        ds_khu_vuc_nho = sorted(kvl.khu_vuc_nho, key=lambda x: x.thu_tu)

        ten_sheet = re.sub(r'[\\/*?:\[\]]', ' ', kvl.ten_khu_vuc).strip()[:31] or f'KhuVuc{kvl.id}'
        ten_goc = ten_sheet
        dem = 1
        while ten_sheet in ten_sheet_da_dung:
            dem += 1
            hau_to = f' ({dem})'
            ten_sheet = ten_goc[:31 - len(hau_to)] + hau_to
        ten_sheet_da_dung.add(ten_sheet)

        ws = wb.create_sheet(title=ten_sheet)
        headers = ['Tên xe', 'Phiên bản'] + [kvn.ten_khu_vuc_nho for kvn in ds_khu_vuc_nho]
        ke_dong_tieu_de(ws, headers)

        for xe in danh_sach_xe:
            gia_map = gia_bl_theo_xe.get(xe.id, {})
            ws.append([xe.ten_xe, xe.phien_ban or ''] + [gia_map.get(kvn.id, 0) or 0 for kvn in ds_khu_vuc_nho])

        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 18
        for idx in range(len(ds_khu_vuc_nho)):
            ws.column_dimensions[get_column_letter(3 + idx)].width = 32

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    vn_time = datetime.now(timezone(timedelta(hours=7)))
    ten_file = f"gia_giay_to_{vn_time.strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=ten_file
    )


@app.route('/admin/api/luu-gia-giay-to-cm', methods=['POST'])
@admin_required
def admin_luu_gia_giay_to_cm():
    """
    Lưu hàng loạt giá giấy tờ Cà Mau (Phường / Xã) cho nhiều xe cùng lúc.
    Body JSON: { "gia": { "<xe_id>": {"phuong": 500000, "xa": 700000}, ... } }
    """
    data = request.get_json(silent=True) or {}
    gia_map = data.get('gia', {}) or {}

    so_luong_cap_nhat = 0
    for xe_id_str, gt in gia_map.items():
        try:
            xe_id = int(xe_id_str)
        except (TypeError, ValueError):
            continue
        if not isinstance(gt, dict):
            continue

        xe = Xe.query.get(xe_id)
        if not xe:
            continue

        gia_phuong_moi = safe_float(gt.get('phuong'), xe.gia_gt_phuong_cm or 0)
        gia_xa_moi = safe_float(gt.get('xa'), xe.gia_gt_xa_cm or 0)

        if (xe.gia_gt_phuong_cm or 0) != gia_phuong_moi or (xe.gia_gt_xa_cm or 0) != gia_xa_moi:
            xe.gia_gt_phuong_cm = gia_phuong_moi
            xe.gia_gt_xa_cm = gia_xa_moi
            so_luong_cap_nhat += 1

    db.session.commit()

    if so_luong_cap_nhat:
        cap_nhat_thoi_gian_dong_bo('Cà Mau', session.get('username'))

    return jsonify({
        'success': True,
        'message': f'Đã cập nhật giá giấy tờ Cà Mau cho {so_luong_cap_nhat} xe.',
        'so_luong': so_luong_cap_nhat
    })


@app.route('/admin/api/luu-gia-giay-to-bl', methods=['POST'])
@admin_required
def admin_luu_gia_giay_to_bl_hang_loat():
    """
    Lưu hàng loạt giá giấy tờ Bạc Liêu theo khu vực nhỏ, cho nhiều xe và cả 3 khu vực lớn cùng lúc.
    Body JSON: { "gia": [ {"xe_id": 1, "khu_vuc_nho_id": 5, "gia": 1200000}, ... ] }
    """
    data = request.get_json(silent=True) or {}
    danh_sach_gia = data.get('gia', []) or []

    so_luong_cap_nhat = 0
    for item in danh_sach_gia:
        if not isinstance(item, dict):
            continue
        try:
            xe_id = int(item.get('xe_id'))
            khu_vuc_nho_id = int(item.get('khu_vuc_nho_id'))
        except (TypeError, ValueError):
            continue
        gia_moi = safe_float(item.get('gia'), 0)

        row = GiaGiayToXeBL.query.filter_by(xe_id=xe_id, khu_vuc_nho_id=khu_vuc_nho_id).first()
        if row:
            if (row.gia or 0) != gia_moi:
                row.gia = gia_moi
                so_luong_cap_nhat += 1
        else:
            db.session.add(GiaGiayToXeBL(xe_id=xe_id, khu_vuc_nho_id=khu_vuc_nho_id, gia=gia_moi))
            so_luong_cap_nhat += 1

    db.session.commit()

    if so_luong_cap_nhat:
        cap_nhat_thoi_gian_dong_bo('Bạc Liêu', session.get('username'))

    return jsonify({
        'success': True,
        'message': f'Đã cập nhật {so_luong_cap_nhat} mức giá giấy tờ Bạc Liêu.',
        'so_luong': so_luong_cap_nhat
    })


@app.route('/api/khu-vuc-giay-to-bl')
def api_khu_vuc_giay_to_bl():
    """Danh sách khu vực lớn + khu vực nhỏ giấy tờ Bạc Liêu (không kèm giá, dùng dựng dropdown)."""
    ds = KhuVucLonBL.query.order_by(KhuVucLonBL.thu_tu).all()
    return jsonify({
        'success': True,
        'khu_vuc_lon': [{
            'id': kvl.id,
            'ma_khu_vuc': kvl.ma_khu_vuc,
            'ten_khu_vuc': kvl.ten_khu_vuc,
            'khu_vuc_nho': [
                {'id': n.id, 'ten_khu_vuc_nho': n.ten_khu_vuc_nho}
                for n in sorted(kvl.khu_vuc_nho, key=lambda x: x.thu_tu)
            ]
        } for kvl in ds]
    })

@app.route('/admin/api/gia-giay-to-bl/<int:xe_id>', methods=['GET'])
@admin_required
def admin_get_gia_giay_to_bl(xe_id):
    """Lấy giá giấy tờ hiện tại của 1 xe theo từng khu vực nhỏ Bạc Liêu (cho form admin)."""
    xe = db.get_or_404(Xe, xe_id)
    return jsonify({'success': True, 'xe_id': xe.id, 'ten_xe': xe.ten_xe,
                     'khu_vuc_lon': lay_gia_giay_to_khu_vuc_nho_bl(xe.id)})

@app.route('/admin/api/gia-giay-to-bl/<int:xe_id>', methods=['POST'])
@admin_required
def admin_update_gia_giay_to_bl(xe_id):
    """
    Cập nhật giá giấy tờ của 1 xe cho các khu vực nhỏ Bạc Liêu.
    Body JSON: { "gia": { "<khu_vuc_nho_id>": 1234000, ... } }
    """
    xe = db.get_or_404(Xe, xe_id)
    data = request.get_json(silent=True) or {}
    gia_map = data.get('gia', {}) or {}

    so_luong_cap_nhat = 0
    for khu_vuc_nho_id_str, gia_moi in gia_map.items():
        try:
            khu_vuc_nho_id = int(khu_vuc_nho_id_str)
        except (TypeError, ValueError):
            continue
        gia_moi = safe_float(gia_moi, 0)

        row = GiaGiayToXeBL.query.filter_by(xe_id=xe.id, khu_vuc_nho_id=khu_vuc_nho_id).first()
        if row:
            if row.gia != gia_moi:
                row.gia = gia_moi
                so_luong_cap_nhat += 1
        else:
            db.session.add(GiaGiayToXeBL(xe_id=xe.id, khu_vuc_nho_id=khu_vuc_nho_id, gia=gia_moi))
            so_luong_cap_nhat += 1

    db.session.commit()

    if so_luong_cap_nhat:
        cap_nhat_thoi_gian_dong_bo('Bạc Liêu', session.get('username'))

    return jsonify({'success': True, 'message': f'Đã cập nhật {so_luong_cap_nhat} khu vực giấy tờ Bạc Liêu cho xe {xe.ten_xe}.'})

@app.route("/admin/update_price/<int:xe_id>", methods=["POST", "PUT"])
@admin_required
def update_regional_price(xe_id):
    xe = db.get_or_404(Xe, xe_id)
    
    data = {}
    if request.form:
        data.update(request.form.to_dict())
    if request.args:
        data.update(request.args.to_dict())
        
    json_data = request.get_json(silent=True)
    if not json_data:
        try:
            if request.data:
                import json
                json_data = json.loads(request.data.decode('utf-8'))
        except Exception:
            pass
    if json_data and isinstance(json_data, dict):
        data.update(json_data)

    if 'gia_bl_cao' in data or 'gia_bl_trung' in data or 'gia_bl_thap' in data:
        vung = 'Bạc Liêu'
    elif 'gia_cm_cao' in data or 'gia_cm_trung' in data or 'gia_cm_thap' in data:
        vung = 'Cà Mau'
    else:
        vung = data.get('vung', '').strip() or session.get('vung', 'Cà Mau')
        
    is_bl = 'bạc liêu' in vung.lower()
    
    old_gia_cao = xe.gia_bl_cao if is_bl else xe.gia_cm_cao
    old_gia_trung = xe.gia_bl_trung if is_bl else xe.gia_cm_trung
    old_gia_thap = xe.gia_bl_thap if is_bl else xe.gia_cm_thap

    if is_bl:
        gia_cao_moi = safe_float(data.get('gia_bl_cao') or data.get('gia_cao'), xe.gia_bl_cao)
        gia_trung_moi = safe_float(data.get('gia_bl_trung') or data.get('gia_trung'), xe.gia_bl_trung)
        gia_thap_moi = safe_float(data.get('gia_bl_thap') or data.get('gia_thap'), xe.gia_bl_thap)
    else:
        gia_cao_moi = safe_float(data.get('gia_cm_cao') or data.get('gia_cao'), xe.gia_cm_cao)
        gia_trung_moi = safe_float(data.get('gia_cm_trung') or data.get('gia_cm_trung'), xe.gia_cm_trung)
        gia_thap_moi = safe_float(data.get('gia_cm_thap') or data.get('gia_thap'), xe.gia_cm_thap)
    
    has_changes = False
    
    if is_bl:
        if xe.gia_bl_cao != gia_cao_moi or xe.gia_bl_trung != gia_trung_moi or xe.gia_bl_thap != gia_thap_moi:
            xe.gia_bl_cao = gia_cao_moi
            xe.gia_bl_trung = gia_trung_moi
            xe.gia_bl_thap = gia_thap_moi
            has_changes = True
    else:
        if xe.gia_cm_cao != gia_cao_moi or xe.gia_cm_trung != gia_trung_moi or xe.gia_cm_thap != gia_thap_moi:
            xe.gia_cm_cao = gia_cao_moi
            xe.gia_cm_trung = gia_trung_moi
            xe.gia_cm_thap = gia_thap_moi
            has_changes = True
            
    db.session.commit()
    
    if has_changes:
        cap_nhat_thoi_gian_dong_bo(vung, session.get('username'))
        
        try:
            sql_log = text("""
                INSERT INTO history_logs (username, action, target_id, old_value, new_value) 
                VALUES (:username, :action, :target_id, :old_val, :new_val)
            """)
            db.session.execute(sql_log, {
                'username': session.get('username', 'Admin'),
                'action': f'Cập nhật giá ({vung})',
                'target_id': f"{xe.ten_xe} (ID: {xe.id})",
                'old_val': f"Cao: {old_gia_cao}, Trung: {old_gia_trung}, Thấp: {old_gia_thap}",
                'new_val': f"Cao: {gia_cao_moi}, Trung: {gia_trung_moi}, Thấp: {gia_thap_moi}"
            })
            db.session.commit()
        except Exception as e:
            print("Lỗi ghi log lịch sử:", e)

    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or data:
        key_name = 'last_updated_bl' if is_bl else 'last_updated_cm'
        user_name = 'last_user_bl' if is_bl else 'last_user_cm'
        st = Setting.query.filter_by(key=key_name).first()
        su = Setting.query.filter_by(key=user_name).first()
        return jsonify({
            "success": True, 
            "message": f"Cập nhật giá {vung} thành công!", 
            "vung": vung,
            "last_updated": st.value if st else "Chưa cập nhật",
            "last_updated_by": su.value if su else "Admin"
        })

@app.route('/admin/history', methods=['GET'])
@admin_required
def admin_history():
    try:
        threshold_date = datetime.now() - timedelta(days=60)
        db.session.execute(
            text("DELETE FROM history_logs WHERE created_at < :threshold"),
            {"threshold": threshold_date}
        )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print("Lỗi tự động xóa lịch sử cũ:", e)

    try:
        sql = text("SELECT * FROM history_logs ORDER BY created_at DESC LIMIT 150")
        result = db.session.execute(sql)
        history_logs = [dict(row._mapping) for row in result]
    except Exception as e:
        return f"""
        <div style="padding: 20px; font-family: sans-serif; color: #721c24; background-color: #f8d7da; border: 1px solid #f5c6cb; border-radius: 5px; margin: 20px;">
            <h3>⚠️ Lỗi cơ sở dữ liệu</h3>
            <p>Có thể bạn chưa chạy lệnh tạo bảng <code>history_logs</code> trong Database.</p>
            <p><b>Chi tiết lỗi:</b> {e}</p>
        </div>
        """, 500

    return render_template('history.html', history_logs=history_logs)
# --- KHỞI TẠO / MIGRATE / SEED DATABASE ---
# Đoạn này chạy VÔ ĐIỀU KIỆN ở mức module (không nằm trong if __name__), vì khi
# deploy bằng Gunicorn (gunicorn app:app), Gunicorn chỉ IMPORT module này chứ
# không bao giờ chạy khối "if __name__ == '__main__':" -> nếu để trong đó thì
# db.create_all(), thêm cột trang_thai, và seed khu vực Bạc Liêu sẽ KHÔNG chạy
# trên Render, dù mọi thứ vẫn chạy bình thường khi test local bằng `python app.py`.
with app.app_context():
    db.create_all()
    # Tự động thêm cột trang_thai nếu cơ sở dữ liệu cũ chưa có
    try:
        db.session.execute(text("ALTER TABLE \"user\" ADD COLUMN IF NOT EXISTS trang_thai VARCHAR(20) DEFAULT 'approved';"))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print("Lỗi tự động thêm cột trang_thai:", e)

    danh_sach_tat_ca_xe = Xe.query.all()
    for xe in danh_sach_tat_ca_xe:
        xe.loai_xe = tu_dong_phan_loai(xe.ten_xe)
    db.session.commit()
    print(f"Đã làm mới phân loại thành công!")

    # Tạo sẵn 3 khu vực lớn + các khu vực nhỏ giấy tờ Bạc Liêu nếu chưa có
    try:
        seed_khu_vuc_giay_to_bl()
        print("Đã đồng bộ khu vực giấy tờ Bạc Liêu!")
    except Exception as e:
        db.session.rollback()
        print("Lỗi khi seed khu vực giấy tờ Bạc Liêu:", e)

# --- KHỞI ĐỘNG ĐỒNG BỘ NỀN ---
# Gọi 1 LẦN duy nhất ở mức module. Trước đây hàm này còn bị gọi thêm 1 lần nữa
# bên trong if __name__ (có kiểm tra WERKZEUG_RUN_MAIN) -> khi chạy `python app.py`
# với debug=True, thực tế nó bị khởi động 2 lần cùng lúc (2 scheduler song song).
# Xoá lệnh gọi trùng, chỉ giữ lại đúng 1 lần ở đây.
#
# LƯU Ý khi lên Render: nếu sau này tăng số worker Gunicorn (--workers > 1), mỗi
# worker là 1 tiến trình riêng và sẽ tự khởi động 1 scheduler riêng -> đồng bộ có
# thể bị chạy trùng nhiều lần cùng lúc. Với 1 worker (mặc định) thì không sao.
start_background_sync()

if __name__ == "__main__":
    # threaded=True: cho phép server xử lý NHIỀU request cùng lúc (ví dụ nhiều người cùng mở
    # ảnh xe, hoặc trình duyệt tải nhiều ảnh màu song song). Mặc định server dev của Flask
    # chỉ xử lý TỪNG request một -> các request ảnh phải xếp hàng chờ nhau, gây cảm giác
    # "chậm" rõ rệt khi có nhiều người dùng cùng lúc trên server thật, dù chạy một mình trên
    # máy cá nhân (localhost) thì không thấy vì không có request nào phải chờ.
    #
    # Khối này CHỈ chạy khi gọi trực tiếp `python app.py` (test local). Khi deploy
    # trên Render bằng Gunicorn, khối này không chạy - Gunicorn tự quản lý việc
    # nhận request, nên không cần app.run() nữa.
    app.run(debug=not IS_PRODUCTION, threaded=True)